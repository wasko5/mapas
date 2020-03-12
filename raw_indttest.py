import pandas as pd
import numpy as np
from scipy import stats
import statsmodels.api as sm
from statsmodels.stats import multitest
import researchpy as rp
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import global_vars
import calculations_helper_functions_only as helper_funcs

#-----------------------------------------------------------0. Input----------------------------------------------------
#TESTING ONLY
'''
global_vars.input_path_and_filename = "Input files\\raw data\\input_raw_independentTtest.xlsx"
global_vars.alpha_threshold = 0.05
global_vars.output_filename = "Test outputs\\output_raw_indttest" #this does not end in .xlsx just for testing; in real app it should end with .xlsx (reason: function will add random numbers to prevent overwriting)

global_vars.input_type = "raw"

global_vars.raw_test = "indttest"

#raw_corr_type = ""
#raw_mr_outcomevar = ""
#raw_mr_predictors = ""
global_vars.raw_indttest_groupvar = "Group"
global_vars.raw_indttest_dv = ["var1", "var2", "var3", "var4", "var5", "var6", "var7", "var8"]
#raw_pairttest_var1 = ""
#raw_pairttest_var2 = ""

#summ_corr_varOne = ""
#summ_corr_varTwo = ""
#summ_corr_coeff = ""
#summ_corr_pvalues = ""

#summ_indttest_var = ""
#summ_indttest_meanOne = ""
#summ_indttest_sdOne = ""
#summ_indttest_nOne = ""
#summ_indttest_meanTwo = ""
#summ_indttest_sdTwo = ""
#summ_indttest_nTwo = ""
#summ_indttest_equal_var = ""

#spss_test = ""
#spss_indttest_nOne = ""
#spss_indttest_nTwo = ""
#spss_indttest_groupOneLabel = ""
#spss_indttest_groupTwoLabel = ""
#spss_pairttest_nOne = ""
#spss_pairttest_nTwo = ""

global_vars.effect_size_choice = "none" #could also be "Hedge's g", "Glass's delta" or blank/none (no idea which one it was)
global_vars.correction_type = "fdr_bh" #see global_vars.master_dict for other values

global_vars.non_numeric_input_raise_errors = True #or False
global_vars.raw_ttest_output_descriptives = True
'''

#-----------------------------------------------------------1. Main flow----------------------------------------------------
#Note that the execution of the main flow is at the bottom
def get_raw_data_df():
	print(global_vars.input_path_and_filename)
	if global_vars.input_path_and_filename.endswith(".xlsx"):
		try:
			raw_data_df = pd.read_excel(global_vars.input_path_and_filename)
		except:
			raise Exception("Oh-oh. For some reason we cannot read the provided file. Please try another file - make sure it's an excel spreadsheet.")

	#raw_data_df.to_excel("Progress dataframes\\raw_indttest\\raw_data_df.xlsx")

	return raw_data_df

def modify_raw_data_df(raw_data_df):
	numeric_cols = global_vars.raw_indttest_dv
	non_numeric_cols = [global_vars.raw_indttest_groupvar] ###list(set(list(raw_data_df.columns)) - set(numeric_cols)) #this is fine because numeric_cols is a subset of df.columns; could be a problem only if numeric_cols could contain unique values - https://stackoverflow.com/questions/3462143/get-difference-between-two-lists
	if global_vars.non_numeric_input_raise_errors == True:
		helper_funcs.error_on_input(df=raw_data_df, cols=numeric_cols, input_type="numeric")
	
	mod_raw_data_df = raw_input_generate_mod_raw_data_df(raw_data_df, numeric_cols, non_numeric_cols)

	#mod_raw_data_df.to_excel("Progress dataframes\\raw_indttest\\mod_raw_data_df.xlsx")

	return mod_raw_data_df

def generate_output_df(mod_raw_data_df):
	output_df = raw_indttest_generate_output_df(mod_raw_data_df)

	#output_df.to_excel("Progress dataframes\\raw_indttest\\output_df.xlsx")

	return output_df

def save_output(mod_raw_data_df, output_df):
	raw_indttest_apa_table(mod_raw_data_df, output_df)
	if global_vars.raw_ttest_output_descriptives == True:
		raw_indttest_descriptives_table(mod_raw_data_df, output_df)

#-----------------------------------------------------------2. Modified raw data dataframe----------------------------------------------------
#2.1.  Main function for generating the modified raw data dataframe
def raw_input_generate_mod_raw_data_df(raw_data_df, numeric_cols, non_numeric_cols=[]):
	mod_raw_data_df = raw_data_df[numeric_cols + non_numeric_cols] #does NOT raise errors when array is blank (default arg)
	mod_raw_data_df[numeric_cols] = mod_raw_data_df[numeric_cols].apply(pd.to_numeric, errors="coerce") #non-numeric values will be np.nan
	try:
		mod_raw_data_df[non_numeric_cols] = mod_raw_data_df[non_numeric_cols].astype(str) #does NOT raise errors when array is blank (default arg)
	except:
		raise Exception("The data could not be processed. Please ensure that the non-numeric columns contain only strings.")
	#mod_raw_data_df = mod_raw_data_df.dropna().reset_index(drop=True) NEVER DOROP NA HERE - JUST COERCE TO KNOW WHERE NA VALUES ARE BUT DO NOT ALTER!!!!
	
	return mod_raw_data_df

#-----------------------------------------------------------3. Output dataframe----------------------------------------------------
#3.2.  Main function for generating the output data dataframe
def raw_indttest_generate_output_df(mod_raw_data_df):
	unique_groups = [x.strip() for x in mod_raw_data_df[global_vars.raw_indttest_groupvar].unique() if len(x.strip()) > 0]
	if len(unique_groups) > 2:
		raise Exception("More than two unique values found in the grouping variable \'{}\'. Please supply data with only two groups.".format(global_vars.raw_indttest_groupvar))
	group1_label, group2_label = unique_groups[0], unique_groups[1] #columns already passed as relevant types so no longer need to control types

	group1_df = mod_raw_data_df[mod_raw_data_df[global_vars.raw_indttest_groupvar]==group1_label]
	group2_df = mod_raw_data_df[mod_raw_data_df[global_vars.raw_indttest_groupvar]==group2_label]

	effect_size_df_row_lookup = {"Cohen's d": 6, "Hedge's g": 7, "Glass's delta": 8}
	dictionaries_list=[]
	for var in global_vars.raw_indttest_dv:

		result = rp.ttest(group1_df[var], group2_df[var], group1_name=group1_label, group2_name=group2_label, 
			 equal_variances=stats.levene(group1_df[var], group2_df[var])[1]>0.05, paired=False)
		ttest_stats_df1 = result[0]
		ttest_stats_df2 = result[1]
		
		current_dict={}
		current_dict["Variable"] = var
		current_dict["All_N"] = int(ttest_stats_df1[ttest_stats_df1["Variable"] == "combined"].iloc[0]["N"])
		current_dict["All_Mean"] = ttest_stats_df1[ttest_stats_df1["Variable"] == "combined"].iloc[0]["Mean"]
		current_dict["All_SD"] = ttest_stats_df1[ttest_stats_df1["Variable"] == "combined"].iloc[0]["SD"]
		current_dict["All_Std Err"] = ttest_stats_df1[ttest_stats_df1["Variable"] == "combined"].iloc[0]["SE"]
		current_dict["All_95% CI_Low"] = ttest_stats_df1[ttest_stats_df1["Variable"] == "combined"].iloc[0]["95% Conf."]
		current_dict["All_95% CI_High"] = ttest_stats_df1[ttest_stats_df1["Variable"] == "combined"].iloc[0]["Interval"]
		current_dict[group1_label+"_N"] = int(ttest_stats_df1[ttest_stats_df1["Variable"] == group1_label].iloc[0]["N"])
		current_dict[group1_label+"_Mean"] = ttest_stats_df1[ttest_stats_df1["Variable"] == group1_label].iloc[0]["Mean"]
		current_dict[group1_label+"_SD"] = ttest_stats_df1[ttest_stats_df1["Variable"] == group1_label].iloc[0]["SD"]
		current_dict[group1_label+"_Std Err"] = ttest_stats_df1[ttest_stats_df1["Variable"] == group1_label].iloc[0]["SE"]
		current_dict[group1_label+"_95% CI_Low"] = ttest_stats_df1[ttest_stats_df1["Variable"] == group1_label].iloc[0]["95% Conf."]
		current_dict[group1_label+"_95% CI_High"] = ttest_stats_df1[ttest_stats_df1["Variable"] == group1_label].iloc[0]["Interval"]
		current_dict[group2_label+"_N"] = int(ttest_stats_df1[ttest_stats_df1["Variable"] == group2_label].iloc[0]["N"])
		current_dict[group2_label+"_Mean"] = ttest_stats_df1[ttest_stats_df1["Variable"] == group2_label].iloc[0]["Mean"]
		current_dict[group2_label+"_SD"] = ttest_stats_df1[ttest_stats_df1["Variable"] == group2_label].iloc[0]["SD"]
		current_dict[group2_label+"_Std Err"] = ttest_stats_df1[ttest_stats_df1["Variable"] == group2_label].iloc[0]["SE"]
		current_dict[group2_label+"_95% CI_Low"] = ttest_stats_df1[ttest_stats_df1["Variable"] == group2_label].iloc[0]["95% Conf."]
		current_dict[group2_label+"_95% CI_High"] = ttest_stats_df1[ttest_stats_df1["Variable"] == group2_label].iloc[0]["Interval"]
		current_dict["Equal Variances Assumed"] = "Yes" if stats.levene(group1_df[var], group2_df[var])[1]>0.05 else "No"
		current_dict["Degrees of Freedom"] = ttest_stats_df2.iloc[1, 1]
		current_dict["t"] = ttest_stats_df2.iloc[2, 1]
		current_dict[global_vars.effect_size_choice] = np.nan if global_vars.effect_size_choice == "none" else ttest_stats_df2.iloc[effect_size_df_row_lookup[global_vars.effect_size_choice], 1]
		current_dict["pvalues"] = ttest_stats_df2.iloc[3,1]

		dictionaries_list.append(current_dict)

	output_df = pd.DataFrame(dictionaries_list)

	return output_df

#-----------------------------------------------------------4. Saving data----------------------------------------------------
#4.2.  Main function for saving data
def raw_indttest_apa_table(mod_raw_data_df, output_df):
	unique_groups = [x.strip() for x in mod_raw_data_df[global_vars.raw_indttest_groupvar].unique() if len(x.strip()) > 0]
	group1_label, group2_label = unique_groups[0], unique_groups[1]

	sign_bool_label = list(output_df.columns)[-1]
	
	apa_output_df = output_df[["Variable","All_Mean","All_SD", group1_label+"_Mean", group1_label+"_SD", group2_label+"_Mean", 
							group2_label+"_SD", "Degrees of Freedom", "t", global_vars.effect_size_choice, "adjusted_pvalues"]]

	#the two operations below are correct so the SettingWithCopyWarning pandas error is supressed temporarily
	pd.options.mode.chained_assignment = None
	apa_output_df[list(apa_output_df.columns)[1:-1]] = apa_output_df[list(apa_output_df.columns)[1:-1]].applymap(lambda x: "{:.2f}".format(x))

	apa_output_df["adjusted_pvalues"] = apa_output_df["adjusted_pvalues"].map(helper_funcs.pvalue_formatting)
	pd.options.mode.chained_assignment = "warn"

	wb = Workbook()
	ws = wb.active
	
	ws.cell(row=1, column=1).value = "Variable"
	ws.merge_cells('A1:A2')
	ws.cell(row=1, column=1).font = global_vars.font_header
	
	ws.cell(row=1, column=2).value = "All, n={}".format(output_df.iloc[0, 1])
	ws.merge_cells('B1:C1')
	
	ws.cell(row=1, column=4).value = "{g}, n={n}".format(g=group1_label, n=output_df.iloc[0,7])
	ws.merge_cells('D1:E1')
	
	ws.cell(row=1, column=6).value = "{g}, n={n}".format(g=group2_label, n=output_df.iloc[0,13])
	ws.merge_cells('F1:G1')
	
	ws.cell(row=1, column=8).value = "df"
	ws.merge_cells('H1:H2')
	ws.cell(row=1, column=8).font = global_vars.font_header
	
	ws.cell(row=1, column=9).value = "t"
	ws.merge_cells('I1:I2')
	ws.cell(row=1, column=9).font = global_vars.font_header
	
	ws.cell(row=1, column=10).value = global_vars.effect_size_choice
	ws.merge_cells('J1:J2')
	ws.cell(row=1, column=10).font = global_vars.font_header
	
	ws.cell(row=1, column=11).value = "p"
	ws.merge_cells('K1:K2')
	ws.cell(row=1, column=11).font = global_vars.font_header

	for col in range(2,7,2):
		ws.cell(row=2, column=col).value = "M"
		ws.cell(row=2, column=col).font = global_vars.font_header
		ws.cell(row=2, column=col+1).value = "SD"
		ws.cell(row=2, column=col+1).font = global_vars.font_header

	for row in dataframe_to_rows(apa_output_df, index=False, header=False):
		ws.append(row)

	for row in range(1, len(apa_output_df)+3):
		for cell in ws[row]:
			cell.alignment = global_vars.alignment_center

	for cell in ws[1]:
		cell.border = Border(top=global_vars.border_APA)
	for cell in ws[2] + ws[len(apa_output_df)+2]:
		cell.border = Border(bottom=global_vars.border_APA)

	if global_vars.effect_size_choice == "none":
		ws.delete_cols(10)

	helper_funcs.add_table_notes(ws, [])

	helper_funcs.save_file("raw_data_indttest", wb)

def raw_indttest_descriptives_table(mod_raw_data_df, output_df):
	if global_vars.correction_type == "none":
		output_df.drop(columns = ["adjusted_pvalues", list(output_df.columns)[-1]], inplace=True)
	
	test_stats_df = output_df[list(output_df.columns)[0:1] + list(output_df.columns)[19:]]
	descriptive_stats_df = output_df.drop(columns = list(output_df.columns)[19:]) #returns what's dropped, not the dropped df

	#the operation below is correct so the SettingWithCopyWarning pandas error is supressed temporarily
	pd.options.mode.chained_assignment = None
	if global_vars.effect_size_choice == "none":
		test_stats_df.drop(columns = [global_vars.effect_size_choice], inplace=True)
	pd.options.mode.chained_assignment = "warn"

	wb = Workbook()
	ws = wb.active
	
	ws.cell(row=1, column=1).value = "Descriptives"
	ws.cell(row=1, column=1).font = global_vars.font_title
	ws.merge_cells('A1:S1')
	
	for row in dataframe_to_rows(descriptive_stats_df, index=False, header=True):
		ws.append(row)
	
	for cell in ws[2]:
		cell.font = global_vars.font_bold
		
	ws.cell(row=len(descriptive_stats_df)+5,column=1).value = "Test Statistics"
	ws.cell(row=len(descriptive_stats_df)+5,column=1).font = global_vars.font_title
	ws.merge_cells(start_row=len(descriptive_stats_df)+5, start_column=1, end_row=len(descriptive_stats_df)+5, end_column=len(test_stats_df.columns))
	
	for row in dataframe_to_rows(test_stats_df, index=False, header=True):
		ws.append(row)

	for cell in ws[len(descriptive_stats_df)+6]:
		cell.font = global_vars.font_bold

	for row in range(1, len(descriptive_stats_df) + len(test_stats_df) + 7):
		for cell in ws[row]:
			cell.alignment = global_vars.alignment_center

	helper_funcs.save_file("raw_data_indttest_descriptives", wb)

def main():
	raw_data_df = get_raw_data_df()
	mod_raw_data_df = modify_raw_data_df(raw_data_df)
	output_df = generate_output_df(mod_raw_data_df)
	output_df = helper_funcs.multitest_correction(output_df)
	save_output(mod_raw_data_df, output_df)