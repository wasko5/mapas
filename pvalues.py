import pandas as pd
import numpy as np
from scipy import stats
import statsmodels.api as sm
from statsmodels.stats import multitest
import researchpy as rp
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import global_vars
import calculations_helper_functions_only as helper_funcs

#-----------------------------------------------------------0. Input----------------------------------------------------
#TESTING ONLY
'''
global_vars.input_path_and_filename = "Input files\\input_p-values.xlsx"
global_vars.alpha_threshold = 0.05
global_vars.output_filename = "Output files for testing\\pvalues" #this does not end in .xlsx just for testing; in real app it should end with .xlsx (reason: function will add random numbers to prevent overwriting)

global_vars.input_type = "pvalues"

global_vars.raw_test = ""

#raw_corr_type = ""
#raw_mr_outcomevar = ""
#raw_mr_predictors = ""
#raw_indttest_groupvar = ""
#raw_indttest_dv = ""
#raw_pairttest_var_pairs = ""

#summ_corr_varOne = ""
#summ_corr_varTwo = ""
#summ_corr_coeff = ""
#summ_corr_pvalues = ""

#summ_indttest_var = ""
#summ_indttest_meanOne = ""
#summ_indttest_sdOne = ""
#summ_indttest_nOne = ""
#summ_indttest_meanTwo = ""
#summ_indttest_sdTwo = ""
#summ_indttest_nTwo = ""
#summ_indttest_equal_var = ""

#spss_test = ""
#spss_indttest_nOne = ""
#spss_indttest_nTwo = ""
#spss_indttest_groupOneLabel = ""
#spss_indttest_groupTwoLabel = ""
#spss_pairttest_nOne = ""
#spss_pairttest_nTwo = ""

#effect_size_choice = "" #could also be "Hedge's g", "Glass's delta" or blank/none (no idea which one it was)
global_vars.correction_type = "fdr_bh" #see global_vars.master_dict for other values

#non_numeric_input_raise_errors = True #defaults to True here
#raw_ttest_output_descriptives = ""
'''

#-----------------------------------------------------------1. Main flow----------------------------------------------------
#Note that the execution of the main flow is at the bottom
'''
def get_raw_data_df():
	if global_vars.input_path_and_filename.endswith(".xlsx"):
		try:
			raw_data_df = pd.read_excel(global_vars.input_path_and_filename)
		except:
			raise Exception("Oh-oh. For some reason we cannot read the provided file. Please try another file - make sure it's an excel spreadsheet.")

	#raw_data_df.to_excel("Progress dataframes\\pvalues\\raw_data_df.xlsx")

	return raw_data_df

def modify_raw_data_df(raw_data_df):
	#-------------------------------------------------------------------TO LET USER CHOOSE PVALUE COLUMN
	mod_raw_data_df = pvalues_generate_mod_raw_data_df(raw_data_df)

	#mod_raw_data_df.to_excel("Progress dataframes\\pvalues\\mod_raw_data_df.xlsx")

	return mod_raw_data_df

def generate_output_df(mod_raw_data_df):
	output_df = pvalues_generate_output_df(mod_raw_data_df)

	#output_df.to_excel("Progress dataframes\\pvalues\\output_df.xlsx")

	return output_df

def save_output(mod_raw_data_df, output_df):
	pvalues_table(mod_raw_data_df, output_df)
'''
#-----------------------------------------------------------2. Modified raw data dataframe----------------------------------------------------
#2.1.  Main function for generating the modified raw data dataframe
def pvalues_generate_mod_raw_data_df(raw_data_df):
	if "pvalues" not in raw_data_df.columns:
		raise Exception("Column \'pvalues\' is not found in the provided file. Please make sure it is typed correctly.")
	helper_funcs.error_on_input(df=raw_data_df, cols=["pvalues"], input_type="numeric")

	mod_raw_data_df = raw_data_df.copy()

	return mod_raw_data_df

#-----------------------------------------------------------3. Output dataframe----------------------------------------------------
#3.2.  Main function for generating the output data dataframe
def pvalues_generate_output_df(mod_raw_data_df):
	output_df = mod_raw_data_df.copy()

	return output_df

#-----------------------------------------------------------4. Saving data----------------------------------------------------
#4.2.  Main function for saving data
def pvalues_table(mod_raw_data_df, output_df):
	wb = Workbook()
	ws = wb.active

	for row in dataframe_to_rows(output_df, index=False, header=True):
		ws.append(row)

	for row in range(1, len(output_df)+2):
		for cell in ws[row]:
			cell.alignment = global_vars.alignment_center

	for cell in ws[1]:
		cell.font = global_vars.font_bold

	for cell in ws[1] + ws[len(output_df)+1]:
		cell.border = Border(bottom=global_vars.border_APA)

		
	helper_funcs.add_table_notes(ws, [])

	#helper_funcs.save_file("pvalues", wb)
	wb.save(filename=global_vars.output_filename + ".xlsx")
'''
def main():
	raw_data_df = get_raw_data_df()
	mod_raw_data_df = modify_raw_data_df(raw_data_df)
	output_df = generate_output_df(mod_raw_data_df)
	output_df = helper_funcs.multitest_correction(output_df)
	save_output(mod_raw_data_df, output_df)
'''