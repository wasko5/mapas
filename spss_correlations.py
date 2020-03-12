import pandas as pd
import numpy as np
from scipy import stats
import statsmodels.api as sm
from statsmodels.stats import multitest
import researchpy as rp
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import global_vars
import calculations_helper_functions_only as helper_funcs

#-----------------------------------------------------------0. Input----------------------------------------------------
#TESTING ONLY
'''
global_vars.input_path_and_filename = "C:\\Users\\Nikolay Petrov\\Desktop\\spearman.xlsx"
#input_fileext = ""
global_vars.alpha_threshold = 0.05
global_vars.output_filename = "Output files for testing\\spss_correlations_" #this does not end in .xlsx just for testing; in real app it should end with .xlsx (reason: function will add random numbers to prevent overwriting)

global_vars.input_type = "spss"

global_vars.raw_test = ""

#raw_corr_type = ""
#raw_mr_outcomevar = ""
#raw_mr_predictors = ""
#raw_indttest_groupvar = ""
#raw_indttest_dv = ""
#raw_pairttest_var_pairs = ""

#summ_corr_varOne = ""
#summ_corr_varTwo = ""
#summ_corr_coeff = ""
#summ_corr_pvalues = ""

#summ_indttest_var = ""
#summ_indttest_meanOne = ""
#summ_indttest_sdOne = ""
#summ_indttest_nOne = ""
#summ_indttest_meanTwo = ""
#summ_indttest_sdTwo = ""
#summ_indttest_nTwo = ""
#summ_indttest_equal_var = ""

global_vars.spss_test = "corr"
#spss_indttest_nOne = ""
#spss_indttest_nTwo = ""
#spss_indttest_groupOneLabel = ""
#spss_indttest_groupTwoLabel = ""
#spss_pairttest_nOne = ""
#spss_pairttest_nTwo = ""

#effect_size_choice = ""
global_vars.correction_type = "fdr_bh" #see global_vars.master_dict for other values

#global_vars.non_numeric_input_raise_errors = True #or False
#raw_ttest_output_descriptives = ""
'''

#-----------------------------------------------------------1. Main flow----------------------------------------------------
#Note that the execution of the main flow is at the bottom
def get_raw_data_df():
	if global_vars.input_path_and_filename.endswith(".xlsx"):
		try:
			raw_data_df = pd.read_excel(global_vars.input_path_and_filename)
		except:
			raise Exception("Oh-oh. For some reason we cannot read the provided file. Please try another file - make sure it's an excel spreadsheet.")

	#raw_data_df.to_excel("Progress dataframes\\spss_correlations\\raw_data_df.xlsx")

	return raw_data_df

def modify_raw_data_df(raw_data_df):
	mod_raw_data_df = spss_corr_generate_mod_raw_data_df(raw_data_df)

	#mod_raw_data_df.to_excel("Progress dataframes\\spss_correlations\\mod_raw_data_df.xlsx")

	return mod_raw_data_df

def generate_output_df(mod_raw_data_df):
	output_df = spss_corr_generate_output_df(mod_raw_data_df)

	#output_df.to_excel("Progress dataframes\\spss_correlations\\output_df.xlsx")

	return output_df

def save_output(mod_raw_data_df, output_df):
	spss_corr_apa_table(mod_raw_data_df, output_df)

#-----------------------------------------------------------2. Modified raw data dataframe----------------------------------------------------
#2.1.  Main function for generating the modified raw data dataframe
def spss_corr_generate_mod_raw_data_df(raw_data_df):
	#-------------------------------------------------MORE RIGOROUS CHECKS HERE--------------------------------------------!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
	mod_raw_data_df = raw_data_df.copy()

	if mod_raw_data_df.iloc[1, 1] == "Pearson Correlation":
		mod_raw_data_df.iloc[0, 0], mod_raw_data_df.iloc[0, 1] = "Variable", "Statistic"
		mod_raw_data_df = mod_raw_data_df.rename(columns=mod_raw_data_df.iloc[0]).drop(mod_raw_data_df.index[0]).reset_index(drop=True)
	elif mod_raw_data_df.iloc[1, 0] == "Kendall's tau_b" or mod_raw_data_df.iloc[1, 0] == "Spearman's rho":
		mod_raw_data_df.iloc[0, 1], mod_raw_data_df.iloc[0, 2] = "Variable", "Statistic"
		mod_raw_data_df = mod_raw_data_df.rename(columns=mod_raw_data_df.iloc[0]).drop(mod_raw_data_df.index[0]).reset_index(drop=True)
		correlation_label = mod_raw_data_df.iloc[0, 0]
		mod_raw_data_df["Statistic"].replace(to_replace="Correlation Coefficient", value=correlation_label, inplace=True)
		mod_raw_data_df.drop(columns = list(mod_raw_data_df.columns)[0], inplace=True)

	return mod_raw_data_df

#-----------------------------------------------------------3. Output dataframe----------------------------------------------------
#3.2.  Main function for generating the output data dataframe
def spss_corr_generate_output_df(mod_raw_data_df):
	'''
	Logic: 1) from the variable column, get the indices of all rows that have a variable 2) loop through those indices and create a standalone dataframe
	for each variable 3) from that dataframe, look at the correlations 4) find the diagonal (correlation is 1) 5) once the diagnoal is found
	then go through each subsequent column and get the needed stats stuff based on the kys of the output_dict
	'''
	correlation_label = mod_raw_data_df.iloc[0, 1]
	output_dict = {"var1": [], "var2": [], correlation_label: [], "pvalues": []}
	vars_row_ind_list = list(mod_raw_data_df[mod_raw_data_df["Variable"].notnull()].index)

	for ind, ele in enumerate(vars_row_ind_list):
		if ind != len(vars_row_ind_list)-1: #check for overflow
			current_df = mod_raw_data_df.iloc[vars_row_ind_list[ind]: vars_row_ind_list[ind+1]]
			row_corrs = current_df.iloc[0, 2:]
			diagonal_flag = False
			for var, corr in row_corrs.items():
				if corr == 1:
					diagonal_flag = True
					continue
				if diagonal_flag == True:
					output_dict["var1"].append(current_df.iloc[0, 0])
					output_dict["var2"].append(var)
					output_dict[correlation_label].append(corr)
					output_dict["pvalues"].append(current_df.loc[ele+1, var])

	output_df = pd.DataFrame(output_dict)

	return output_df

#-----------------------------------------------------------4. Saving data----------------------------------------------------
#4.2.  Main function for saving data
def spss_corr_apa_table(mod_raw_data_df, output_df):
	#could very easily use the summ_corr_apa_table function here as I pass identical data - seperate is preferred, however, as I might want to adjust styling or sth
	#note that the raw corr apa table has CIs so not usable
	correlation_label = mod_raw_data_df.iloc[0, 1]
	variables_list = list(mod_raw_data_df.columns)[2:]

	wb = Workbook()
	ws = wb.active

	ws.append([""] + variables_list[1:])
	for ind,var in enumerate(variables_list[:-1]):
		ws.cell(row=ind+2, column=1).value = var

	start_col = 2
	for row in range(2, len(variables_list)+1):
		row_var = ws.cell(row=row, column=1).value
		for col in range(start_col, len(variables_list)+1):
			col_var = ws.cell(row=1, column=col).value
			#here query method is not only slower as it is much smaller dataset but also cannot refer to two different vaiables (colname and val)
			df_filter = output_df[((output_df["var1"]==row_var) & (output_df["var2"]==col_var)) | ((output_df["var1"]==col_var) & (output_df["var2"]==row_var))].iloc[0]
			r = df_filter[correlation_label]
			p = df_filter["adjusted_pvalues"]
			r = helper_funcs.correlations_format_val(r, p)
			ws.cell(row=row, column=col).value = r
		start_col += 1

	for row in range(1, len(variables_list)+1):
		for cell in ws[row]:
			cell.alignment = global_vars.alignment_center

	for cell in ws[1]:
		cell.font = global_vars.font_header

	for row in range(2, len(variables_list)+1):
		ws.cell(row=row, column=1).font = global_vars.font_header

	for cell in ws[len(variables_list)]:
		cell.border = Border(bottom=global_vars.border_APA)

	for cell in ws[1]:
		cell.border = Border(top=global_vars.border_APA, bottom=global_vars.border_APA)

	table_notes = ["**p < 0.01", "*p < {}".format(global_vars.alpha_threshold)]
	helper_funcs.add_table_notes(ws, table_notes)

	helper_funcs.save_file("spss_correlations", wb)




	'''

	correlation_label = output_df.columns[2]
	variables_list = mod_raw_data_df.columns[1:]

	apa_table_df = pd.DataFrame(index = variables_list[:-1], columns = variables_list[1:])

	column = 0
	for row in range(0, len(apa_table_df)):
		row_var = apa_table_df.index[row]
		for col in range(column, len(apa_table_df.columns)):
			col_var = apa_table_df.columns[col]
			r = output_df[((output_df["var1"]==row_var) & (output_df["var2"]==col_var)) | ((output_df["var1"]==col_var) & (output_df["var2"]==row_var))][correlation_label].item()
			p = output_df[((output_df["var1"]==row_var) & (output_df["var2"]==col_var)) | ((output_df["var1"]==col_var) & (output_df["var2"]==row_var))]["adjusted_pvalues"].item()
			apa_table_df.iloc[row, col] = correlations_format_val(r, p)
		column += 1

	wb = Workbook()
	ws = wb.active

	for row in dataframe_to_rows(apa_table_df, index=True, header=True):
		ws.append(row)
	ws.delete_rows(2) #the second row in the spreadsheet is blank for some reason so we get rid of it

	for cell in ws['A'] + ws[1]:
		cell.font = font_header
		
	for col in range(1, len(apa_table_df.columns)+2):
		for row in range(1, len(apa_table_df)+2):
			ws.cell(row=row, column=col).alignment = alignment_center
	
	for cell in ws[1]:
		cell.border = Border(top=border_APA, bottom=border_APA)

	for cell in ws[len(apa_table_df)+1]:
		cell.border = Border(bottom=border_APA)

	table_notes = ["Correlation coefficient used: {}".format(correlation_label), "**p < 0.01", "*p < {}".format(alpha_threshold)]
	add_table_notes(ws, table_notes)

	save_file("spss_correlations", wb)
'''

def main():
	raw_data_df = get_raw_data_df()
	mod_raw_data_df = modify_raw_data_df(raw_data_df)
	output_df = generate_output_df(mod_raw_data_df)
	output_df = helper_funcs.multitest_correction(output_df)
	save_output(mod_raw_data_df, output_df)