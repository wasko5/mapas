import pandas as pd
import numpy as np
from scipy import stats
import statsmodels.api as sm
from statsmodels.stats import multitest
import researchpy as rp
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import global_vars
import calculations_helper_functions_only as helper_funcs

#-----------------------------------------------------------0. Input----------------------------------------------------
#TESTING ONLY
'''
global_vars.input_path_and_filename = "Input files\\input_summary_independettest.xlsx"
#input_fileext = ""
global_vars.alpha_threshold = 0.05
global_vars.output_filename = "Output files for testing\\summ_indttest" #this does not end in .xlsx just for testing; in real app it should end with .xlsx (reason: function will add random numbers to prevent overwriting)

global_vars.input_type = "summ_indttest"

global_vars.raw_test = ""

#raw_corr_type = ""
#raw_mr_outcomevar = ""
#raw_mr_predictors = ""
#raw_indttest_groupvar = ""
#raw_indttest_dv = ""
#raw_pairttest_var_pairs = ""

#summ_corr_varOne = ""
#summ_corr_varTwo = ""
#summ_corr_coeff = ""
#summ_corr_pvalues = ""

global_vars.summ_indttest_var = "Variable"
global_vars.summ_indttest_meanOne = "Mean1"
global_vars.summ_indttest_sdOne = "SD1"
global_vars.summ_indttest_nOne = "N1"
global_vars.summ_indttest_meanTwo = "Mean2"
global_vars.summ_indttest_sdTwo = "SD2"
global_vars.summ_indttest_nTwo = "N2"
global_vars.summ_indttest_equal_var = ""

#spss_test = ""
#spss_indttest_nOne = ""
#spss_indttest_nTwo = ""
#spss_indttest_groupOneLabel = ""
#spss_indttest_groupTwoLabel = ""
#spss_pairttest_nOne = ""
#spss_pairttest_nTwo = ""

global_vars.effect_size_choice = "none" #could also be "Hedge's g", "Glass's delta" or blank/none (no idea which one it was)
global_vars.correction_type = "none" #see global_vars.master_dict for other values

#non_numeric_input_raise_errors = True #defaults to True here
#raw_ttest_output_descriptives = ""
'''

#-----------------------------------------------------------1. Main flow----------------------------------------------------
#Note that the execution of the main flow is at the bottom
'''
def get_raw_data_df():
	if global_vars.input_path_and_filename.endswith(".xlsx"):
		try:
			raw_data_df = pd.read_excel(global_vars.input_path_and_filename)
		except:
			raise Exception("Oh-oh. For some reason we cannot read the provided file. Please try another file - make sure it's an excel spreadsheet.")

	#raw_data_df.to_excel("Progress dataframes\\summ_indttest\\raw_data_df.xlsx")

	return raw_data_df

def modify_raw_data_df(raw_data_df):
	numeric_cols = [global_vars.summ_indttest_meanOne, global_vars.summ_indttest_sdOne, global_vars.summ_indttest_nOne, global_vars.summ_indttest_meanTwo, global_vars.summ_indttest_sdTwo, global_vars.summ_indttest_nTwo]
	non_numeric_cols = [global_vars.summ_indttest_var]
	if global_vars.summ_indttest_equal_var != "":
		non_numeric_cols.append(global_vars.summ_indttest_equal_var)
	helper_funcs.error_on_input(df=raw_data_df, cols=numeric_cols, input_type="numeric")
	helper_funcs.error_on_input(df=raw_data_df, cols=non_numeric_cols, input_type="nan")
	mod_raw_data_df = summ_inddtest_generate_mod_raw_data_df(raw_data_df, numeric_cols, non_numeric_cols)

	#mod_raw_data_df.to_excel("Progress dataframes\\summ_indttest\\mod_raw_data_df.xlsx")

	return mod_raw_data_df

def generate_output_df(mod_raw_data_df):
	output_df = summ_indttest_generate_output_df(mod_raw_data_df)

	#output_df.to_excel("Progress dataframes\\summ_indttest\\output_df.xlsx")

	return output_df

def save_output(mod_raw_data_df, output_df):
	summ_indttest_apa_table(mod_raw_data_df, output_df)
'''
#-----------------------------------------------------------2. Modified raw data dataframe----------------------------------------------------
#2.1.  Main function for generating the modified raw data dataframe
def summ_inddtest_generate_mod_raw_data_df(raw_data_df, numeric_cols, non_numeric_cols=[]):
	mod_raw_data_df = raw_data_df[numeric_cols + non_numeric_cols] #does NOT raise errors when array is blank (default arg)
	mod_raw_data_df[numeric_cols] = mod_raw_data_df[numeric_cols].apply(pd.to_numeric, errors="coerce") #non-numeric values will be np.nan
	try:
		mod_raw_data_df[global_vars.summ_indttest_var] = mod_raw_data_df[global_vars.summ_indttest_var].astype(str) #does NOT raise errors when array is blank (default arg)
	except:
		raise Exception("The data could not be processed. Please ensure that the non-numeric columns contain only strings.")
	
	if global_vars.summ_indttest_equal_var == "":
		mod_raw_data_df["Equal Variances"] = [True] * len(mod_raw_data_df)
	else:
		if not raw_data_df[global_vars.summ_indttest_equal_var].isin([True, False]).all():
			raise Exception("A value different from True or False is found in column \'{}\'.\nPlease ensure that only True and False values are provided.".format(global_vars.summ_indttest_equal_var))
		
	return mod_raw_data_df

#-----------------------------------------------------------3. Output dataframe----------------------------------------------------
#3.2.  Main function for generating the output data dataframe
def summ_indttest_generate_output_df(mod_raw_data_df):
	if global_vars.summ_indttest_equal_var != "":
		equal_var_col = global_vars.summ_indttest_equal_var
	else:
		equal_var_col = list(mod_raw_data_df.columns)[-1]

	output_dict = {global_vars.summ_indttest_var: [], global_vars.summ_indttest_meanOne: [], global_vars.summ_indttest_sdOne: [], global_vars.summ_indttest_meanTwo: [],
					global_vars.summ_indttest_sdTwo: [], "Degrees of Freedom": [], "t": [], global_vars.effect_size_choice: [], "pvalues": []}

	for row in range(0, len(mod_raw_data_df)):
		current_series = mod_raw_data_df.loc[row]
		n1 = current_series[global_vars.summ_indttest_nOne]
		n2 = current_series[global_vars.summ_indttest_nTwo]

		output_dict[global_vars.summ_indttest_var].append(current_series[global_vars.summ_indttest_var])
		output_dict[global_vars.summ_indttest_meanOne].append(current_series[global_vars.summ_indttest_meanOne])
		output_dict[global_vars.summ_indttest_sdOne].append(current_series[global_vars.summ_indttest_sdOne])
		output_dict[global_vars.summ_indttest_meanTwo].append(current_series[global_vars.summ_indttest_meanTwo])
		output_dict[global_vars.summ_indttest_sdTwo].append(current_series[global_vars.summ_indttest_sdTwo])

		if current_series[equal_var_col]: #equal vars assumed
			deg_free = n1 + n2 - 2
		else:
			s1_pooled = current_series[global_vars.summ_indttest_sdOne]**2
			s2_pooled = current_series[global_vars.summ_indttest_sdTwo]**2
			deg_free = (s1_pooled/n1 + s2_pooled/n2)**2 / ( (1/(n1-1)*(s1_pooled/n1)**2) + (1/(n2-1)*(s2_pooled/n2)**2) )
		output_dict["Degrees of Freedom"].append(deg_free)
		#stats.ttest_ind_from_stats itself decides whether to to perform student's t-test or welch's based on equal_var
		t, p = stats.ttest_ind_from_stats(mean1 = current_series[global_vars.summ_indttest_meanOne], std1 = current_series[global_vars.summ_indttest_sdOne], 
										nobs1 = n1, mean2 = current_series[global_vars.summ_indttest_meanTwo], std2 = current_series[global_vars.summ_indttest_sdTwo], 
										nobs2 = n2, equal_var=current_series[equal_var_col])
		output_dict["t"].append(t)
		output_dict["pvalues"].append(p)
		
		if global_vars.effect_size_choice == "Cohen's d":
			effect_size = t*np.sqrt(1/n1 + 1/n2)
		elif global_vars.effect_size_choice == "Hedge's g":
			#calculated using cohens_d * (1 - (3/4(n1+n2-9))) where cohens d is t*sqrt(1/n1 + 1/n2)
			effect_size = (t*np.sqrt(1/n1 + 1/n2))*np.sqrt(1/n1 + 1/n2)
		elif global_vars.effect_size_choice == "Glass's delta":
			#glass delta = (Mean2 - Mean1) / SD1 where SD1 is always that of control
			effect_size = (current_series[summ_indttest_meanTwo] - current_series[summ_indttest_meanOne]) / current_series[summ_indttest_sdOne]
		elif global_vars.effect_size_choice == "None":
			effect_size = np.nan
		output_dict[global_vars.effect_size_choice].append(effect_size)

	output_df = pd.DataFrame(output_dict)
	
	return output_df

#-----------------------------------------------------------4. Saving data----------------------------------------------------
#4.2.  Main function for saving data
def summ_indttest_apa_table(mod_raw_data_df, output_df):
	'''
	sign_bool_label = list(output_df.columns)[-1]

	if output_pvalues_type == "adjusted":
		adjusted_pvalues_threshold = None
		output_df["pvalues"] = output_df["adjusted_pvalues"]
	else:
		adjusted_pvalues_threshold = sign_bool_label[sign_bool_label.find("=")+2:]
	output_df.drop(columns = ["adjusted_pvalues", sign_bool_label], inplace=True)
	'''
	
	#output_df.drop(columns = ["pvalues", list(output_df.columns)[-1]], inplace=True)
	output_df.drop(columns = ["pvalues"], inplace=True)
	print(output_df)

	pd.options.mode.chained_assignment = None
	output_df[list(output_df.columns)[1:-1]] = output_df[list(output_df.columns)[1:-1]].applymap(lambda x: "{:.2f}".format(x))

	output_df["adjusted_pvalues"] = output_df["adjusted_pvalues"].map(helper_funcs.pvalue_formatting)
	pd.options.mode.chained_assignment = "warn"

	wb = Workbook()
	ws = wb.active

	ws.cell(row=1, column=1).value = "Variable"
	ws.merge_cells("A1:A2")
	ws.cell(row=1, column=1).font = global_vars.font_header
	
	ws.cell(row=1, column=2).value = "Group1, n={n}".format(n=mod_raw_data_df[global_vars.summ_indttest_nOne][0])
	ws.merge_cells("B1:C1")
	
	ws.cell(row=1, column=4).value = "Group2, n={n}".format(n=mod_raw_data_df[global_vars.summ_indttest_nTwo][0])
	ws.merge_cells("D1:E1")
	
	ws.cell(row=1, column=6).value = "df"
	ws.merge_cells("F1:F2")
	ws.cell(row=1, column=6).font = global_vars.font_header
	
	ws.cell(row=1, column=7).value = "t"
	ws.merge_cells("G1:G2")
	ws.cell(row=1, column=7).font = global_vars.font_header
	
	ws.cell(row=1, column=8).value = global_vars.effect_size_choice
	ws.merge_cells("H1:H2")
	ws.cell(row=1, column=8).font = global_vars.font_header
	
	ws.cell(row=1, column=9).value = "p"
	ws.merge_cells("I1:I2")
	ws.cell(row=1, column=9).font = global_vars.font_header
	
	for col in range(2,5,2):
		ws.cell(row=2, column=col).value = "M"
		ws.cell(row=2, column=col).font = global_vars.font_header
		ws.cell(row=2, column=col+1).value = "SD"
		ws.cell(row=2, column=col+1).font = global_vars.font_header
	
	for row in dataframe_to_rows(output_df, index=False, header=False):
		ws.append(row)

	for row in range(1, len(output_df)+3):
		for cell in ws[row]:
			cell.alignment = global_vars.alignment_center

	for cell in ws[1]:
		cell.border = Border(top=global_vars.border_APA)
	for cell in ws[2] + ws[len(output_df)+2]:
		cell.border = Border(bottom=global_vars.border_APA)

	if global_vars.effect_size_choice == "None":
		ws.delete_cols(8)

	helper_funcs.add_table_notes(ws, [])

	#helper_funcs.save_file("summ_indttest", wb)
	wb.save(filename=global_vars.output_filename + ".xlsx")

'''
def main():
	raw_data_df = get_raw_data_df()
	mod_raw_data_df = modify_raw_data_df(raw_data_df)
	output_df = generate_output_df(mod_raw_data_df)
	output_df = helper_funcs.multitest_correction(output_df)
	save_output(mod_raw_data_df, output_df)
'''