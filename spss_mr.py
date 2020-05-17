import pandas as pd
import numpy as np
from scipy import stats
import statsmodels.api as sm
from statsmodels.stats import multitest
import researchpy as rp
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import global_vars
import calculations_helper_functions_only as helper_funcs

#-----------------------------------------------------------0. Input----------------------------------------------------
#TESTING ONLY
'''
global_vars.input_path_and_filename = "Input files\\spss\\input_spss_mr_all stats requested.xlsx"
global_vars.alpha_threshold = 0.05
global_vars.output_filename = "Output files for testing\\spss_mr" #this does not end in .xlsx just for testing; in real app it should end with .xlsx (reason: function will add random numbers to prevent overwriting)

global_vars.input_type = "spss"

global_vars.raw_test = ""

#raw_corr_type = ""
#raw_mr_outcomevar = ""
#raw_mr_predictors = ""
#raw_indttest_groupvar = ""
#raw_indttest_dv = ""
#raw_pairttest_var_pairs = ""

#summ_corr_varOne = ""
#summ_corr_varTwo = ""
#summ_corr_coeff = ""
#summ_corr_pvalues = ""

#summ_indttest_var = ""
#summ_indttest_meanOne = ""
#summ_indttest_sdOne = ""
#summ_indttest_nOne = ""
#summ_indttest_meanTwo = ""
#summ_indttest_sdTwo = ""
#summ_indttest_nTwo = ""
#summ_indttest_equal_var = ""

global_vars.spss_test = "mr"
#spss_indttest_nOne = ""
#spss_indttest_nTwo = ""
#spss_indttest_groupOneLabel = ""
#spss_indttest_groupTwoLabel = ""
#spss_pairttest_nOne = ""
#spss_pairttest_nTwo = ""

#effect_size_choice = ""
global_vars.correction_type = "none" #see global_vars.master_dict for other values

#global_vars.non_numeric_input_raise_errors = True #or False
#raw_ttest_output_descriptives = ""
'''

#-----------------------------------------------------------1. Main flow----------------------------------------------------
#Note that the execution of the main flow is at the bottom
'''
def get_raw_data_df():
	if global_vars.input_path_and_filename.endswith(".xlsx"):
		try:
			raw_data_df = pd.read_excel(global_vars.input_path_and_filename)
		except:
			raise Exception("Oh-oh. For some reason we cannot read the provided file. Please try another file - make sure it's an excel spreadsheet.")

	return raw_data_df

def modify_raw_data_df(raw_data_df):
	mod_raw_data_df = spss_mr_generate_mod_raw_data_df(raw_data_df) #might remove the function depending on whether I add logical checks

	#mod_raw_data_df.to_excel("Progress dataframes\\spss_mr\\mod_raw_data_df.xlsx")

	return mod_raw_data_df

def generate_output_df(mod_raw_data_df):
	output_df = spss_mr_generate_output_df(mod_raw_data_df)

	#output_df.to_excel("Progress dataframes\\spss_mr\\output_df.xlsx")

	return output_df

def save_output(mod_raw_data_df, output_df):
	spss_mr_apa_table(mod_raw_data_df, output_df)
'''
#-----------------------------------------------------------2. Modified raw data dataframe----------------------------------------------------
#2.1.  Main function for generating the modified raw data dataframe
def spss_mr_generate_mod_raw_data_df(raw_data_df):
	if not(raw_data_df.columns[0] == "Coefficientsa" or
		raw_data_df.columns[1:].str.contains("Unnamed:").all() or
		list(raw_data_df.iloc[0, 0:7]) == ["Model", np.nan, "Unstandardized Coefficients", np.nan, "Standardized Coefficients", "t", "Sig."] or
		list(raw_data_df.iloc[1, 0:7]) == [np.nan, np.nan, "B", "Std. Error", "Beta", np.nan, np.nan] or
		"Dependent Variable:" in raw_data_df.iloc[raw_data_df.index[-1], 0] or
		raw_data_df.iloc[2, 1] == "(Constant)"):
		
		raise Exception("The provided SPSS regression table does not seem to be in the accepted format. Please use the \"How to export SPSS table?\" button for guidance or refer to the documentation.")

	mod_raw_data_df = raw_data_df.copy()

	#renaming columns as the multi-level formatting of the spss table makes it difficult to work with
	renaming_dict = {}
	for ind, col_name in enumerate(mod_raw_data_df.columns):
		if "Unnamed" in col_name:
			if not pd.isna(mod_raw_data_df.iloc[0, ind]):
				renaming_dict[col_name] = mod_raw_data_df.iloc[0, ind]
			elif not pd.isna(mod_raw_data_df.iloc[1, ind]):
				renaming_dict[col_name] = mod_raw_data_df.iloc[1, ind]

	mod_raw_data_df = mod_raw_data_df.rename(columns=renaming_dict).drop([mod_raw_data_df.index[0], mod_raw_data_df.index[1]]).reset_index(drop=True)

	return mod_raw_data_df

#-----------------------------------------------------------3. Output dataframe----------------------------------------------------
#3.2.  Main function for generating the output data dataframe
def spss_mr_generate_output_df(mod_raw_data_df):
	variables_list = ["(Constant)"] + list(mod_raw_data_df[mod_raw_data_df.columns[1]])[1:-1]

	try:
		ci_low_list = ["{:.2f}".format(x) for x in list(mod_raw_data_df["95.0% Confidence Interval for B"])[:-1]]
	except KeyError:
		ci_low_list = [""] * (len(variables_list))
	try:
		ci_high_list = ["{:.2f}".format(x) for x in list(mod_raw_data_df["Upper Bound"])[:-1]]
	except KeyError:
		ci_high_list = [""] * (len(variables_list))
	beta_list = ["{:.2f}".format(x) for x in list(mod_raw_data_df["Standardized Coefficients"])[:-1]]

	output_df = pd.DataFrame()
	output_df["Variable"] = variables_list
	output_df["B"] = ["{:.2f}".format(x) for x in list(mod_raw_data_df["Unstandardized Coefficients"])[:-1]]
	output_df["95% CI"] = ["["+low+","+high+"]" for low,high in zip(ci_low_list, ci_high_list)]
	output_df["beta"] = ['' if x=='nan' else x for x in beta_list] #the beta for the constant is missing from the spss output
	output_df["t"] = ["{:.2f}".format(x) for x in list(mod_raw_data_df["t"])[:-1]]
	output_df["p"] = [helper_funcs.pvalue_formatting(x) for x in list(mod_raw_data_df["Sig."])[:-1]] #can afford to format here as there will be no multitest corrections applied
	
	return output_df

#-----------------------------------------------------------4. Saving data----------------------------------------------------
#4.2.  Main function for saving data
def spss_mr_apa_table(mod_raw_data_df, output_df):
	wb = Workbook()
	ws = wb.active

	for row in dataframe_to_rows(output_df, index=False, header=True):
		ws.append(row)

	for cell in ws[1]:
		cell.font = global_vars.font_header
		
	for row in range(1, len(output_df)+2):
		for cell in ws[row]:
			cell.alignment = global_vars.alignment_center
	
	for cell in ws[1]:
		cell.border = Border(bottom=global_vars.border_APA, top=global_vars.border_APA)

	for cell in ws[len(output_df)+1]:
		cell.border = Border(bottom=global_vars.border_APA)

	DV_cell = mod_raw_data_df[mod_raw_data_df.columns[0]][len(mod_raw_data_df[mod_raw_data_df.columns[0]])-1]
	DV = DV_cell[DV_cell.find(":")+2:]
	table_notes = ["R squared adjusted = X.XX", "Dependent variable: {}".format(DV)]
	if output_df["95% CI"][0] == "[,]":
		table_notes.append("Confidence intervals were not found in the SPSS table. Please add them to your SPSS table and re-run the program or add them manually.")

	helper_funcs.add_table_notes(ws, table_notes)

	#helper_funcs.save_file("spss_mr", wb)
	wb.save(filename=global_vars.output_filename + ".xlsx")
'''
def main():
	raw_data_df = get_raw_data_df()
	mod_raw_data_df = modify_raw_data_df(raw_data_df)
	output_df = generate_output_df(mod_raw_data_df)
	output_df = helper_funcs.multitest_correction(output_df)
	save_output(mod_raw_data_df, output_df)
'''