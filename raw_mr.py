import pandas as pd
import numpy as np
from scipy import stats
import statsmodels.api as sm
#from statsmodels.stats import multitest
import researchpy as rp
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import global_vars
import calculations_helper_functions_only as helper_funcs

#-----------------------------------------------------------0. Input----------------------------------------------------
#TESTING ONLY
'''
global_vars.input_path_and_filename = "Input files\\raw data\\input_raw_mr.xlsx"
global_vars.alpha_threshold = 0.05
global_vars.output_filename = "Output files for testing\\raw_mr" #this does not end in .xlsx just for testing; in real app it should end with .xlsx (reason: function will add random numbers to prevent overwriting)

global_vars.input_type = "raw"

global_vars.raw_test = "mr"

#raw_corr_type = "pearson" #could be "spearman" or "kendall"
global_vars.raw_mr_outcomevar = "var5"
global_vars.raw_mr_predictors = ["var1", "var2", "var3", "var4"] #not used currently but should be used ASAP
#raw_indttest_groupvar = ""
#raw_pairttest_var1 = ""
#raw_pairttest_var2 = ""

#summ_corr_varOne = ""
#summ_corr_varTwo = ""
#summ_corr_coeff = ""
#summ_corr_pvalues = ""

#summ_indttest_var = ""
#summ_indttest_meanOne = ""
#summ_indttest_sdOne = ""
#summ_indttest_nOne = ""
#summ_indttest_meanTwo = ""
#summ_indttest_sdTwo = ""
#summ_indttest_nTwo = ""
#summ_indttest_equal_var = ""

#spss_test = ""
#spss_indttest_nOne = ""
#spss_indttest_nTwo = ""
#spss_indttest_groupOneLabel = ""
#spss_indttest_groupTwoLabel = ""
#spss_pairttest_nOne = ""
#spss_pairttest_nTwo = ""

#effect_size_choice = ""
global_vars.correction_type = "none" #see global_vars.master_dict for other values

global_vars.non_numeric_input_raise_errors = True #or False
#raw_ttest_output_descriptives = ""
'''

#-----------------------------------------------------------1. Main flow----------------------------------------------------
#Note that the execution of the main flow is at the bottom
'''
def get_raw_data_df():
	if global_vars.input_path_and_filename.endswith(".xlsx"):
		try:
			raw_data_df = pd.read_excel(global_vars.input_path_and_filename)
		except:
			raise Exception("Oh-oh. For some reason we cannot read the provided file. Please try another file - make sure it's an excel spreadsheet.")

	#raw_data_df.to_excel("Progress dataframes\\raw_mr\\raw_data_df.xlsx")

	return raw_data_df

def modify_raw_data_df(raw_data_df):
	numeric_cols = [global_vars.raw_mr_outcomevar] + global_vars.raw_mr_predictors
	if global_vars.non_numeric_input_raise_errors == True:
		helper_funcs.error_on_input(df=raw_data_df, cols=numeric_cols, input_type="numeric")
	
	mod_raw_data_df = raw_input_generate_mod_raw_data_df(raw_data_df, numeric_cols)

	#mod_raw_data_df.to_excel("Progress dataframes\\raw_mr\\mod_raw_data_df.xlsx")

	return mod_raw_data_df

def generate_output_df(mod_raw_data_df):
	output_df = raw_mr_generate_output_df(mod_raw_data_df)

	#output_df.to_excel("Progress dataframes\\raw_mr\\output_df.xlsx")

	return output_df

def save_output(mod_raw_data_df, output_df):
	raw_mr_apa_table(mod_raw_data_df, output_df)
'''
#-----------------------------------------------------------2. Modified raw data dataframe----------------------------------------------------
#2.1.  Main function for generating the modified raw data dataframe
'''
def raw_input_generate_mod_raw_data_df(raw_data_df, numeric_cols, non_numeric_cols=[]):
	mod_raw_data_df = raw_data_df[numeric_cols + non_numeric_cols] #does NOT raise errors when array is blank (default arg)
	mod_raw_data_df[numeric_cols] = mod_raw_data_df[numeric_cols].apply(pd.to_numeric, errors="coerce") #non-numeric values will be np.nan
	try:
		mod_raw_data_df[non_numeric_cols] = mod_raw_data_df[non_numeric_cols].astype(str) #does NOT raise errors when array is blank (default arg)
	except:
		raise Exception("The data could not be processed. Please ensure that the non-numeric columns contain only strings.")
	#mod_raw_data_df = mod_raw_data_df.dropna().reset_index(drop=True) NEVER DOROP NA HERE - JUST COERCE TO KNOW WHERE NA VALUES ARE BUT DO NOT ALTER!!!!
	
	return mod_raw_data_df
'''
#-----------------------------------------------------------3. Output dataframe----------------------------------------------------
#3.2.  Main function for generating the output data dataframe
def raw_mr_generate_output_df(mod_raw_data_df):	
	formula = global_vars.raw_mr_outcomevar + "~"
	for var in global_vars.raw_mr_predictors:
		formula = formula + var + "+"
	formula = formula[:-1] #the loop leaves a trailing "+" sign so this gets rid of it

	result = sm.OLS.from_formula(formula=formula, data=mod_raw_data_df).fit() #everything before the .fit() is model specification
	
	#standardization of the raw data dataframe so it can also produce beta values for the output
	mod_raw_data_df_z = mod_raw_data_df.apply(lambda x: stats.zscore(x, nan_policy="omit"))
	result_z = sm.OLS.from_formula(formula=formula, data=mod_raw_data_df_z).fit()
	beta_list = list(result_z.params)

	CI_B_list = []
	for index, row in result.conf_int().iterrows(): #weirdly enough result_conf_int() returns a dataframe with the CIs
		CI_B_list.append(list(row))
		
	CI_beta_list = []
	for index, row in result_z.conf_int().iterrows():
		CI_beta_list.append(list(row))

	output_dict = {}
	output_dict["Variable"] = ["(Constant)"] + global_vars.raw_mr_predictors
	output_dict["B"] = list(result.params)
	output_dict["Std Err B"] = list(result.bse)
	output_dict["95% CI B"] = ["[" + ",".join("{:.2f}".format(e) for e in pair) + "]" for pair in CI_B_list]
	output_dict["beta"] = beta_list
	output_dict["Std Err beta"] = list(result_z.bse) #might not need - if so, amend the apa_table func
	output_dict["95% CI beta"] = ["[" + ",".join("{:.2f}".format(e) for e in pair) + "]" for pair in CI_beta_list] #might not need
	output_dict["t"] = list(result.tvalues)
	output_dict["pvalues"] = list(result.pvalues)
	output_dict["R2"] = [result.rsquared] + [np.nan] * (len(output_dict["Variable"])-1)
	output_dict["R2adj"] = [result.rsquared_adj] + [np.nan] * (len(output_dict["Variable"])-1)

	output_df = pd.DataFrame(output_dict)
	
	return output_df

#-----------------------------------------------------------4. Saving data----------------------------------------------------
#4.2.  Main function for saving data
def raw_mr_apa_table(mod_raw_data_df, output_df):
	mod_output_df = output_df[["Variable","B","95% CI B","beta","t","pvalues"]]

	pd.options.mode.chained_assignment = None
	mod_output_df[["B", "beta", "t"]] = mod_output_df[["B", "beta", "t"]].applymap(lambda x: "{:.2f}".format(x))

	mod_output_df["pvalues"] = mod_output_df["pvalues"].map(helper_funcs.pvalue_formatting)
	pd.options.mode.chained_assignment = "warn"

	mod_output_df.rename(columns = {"95% CI B": "95% CI", "pvalues": "p"}, inplace=True)
	mod_output_df.loc[0, "beta"] = "" #removes the beta value for constant as it is always 0

	wb = Workbook()
	ws = wb.active

	ws.append(list(mod_output_df.columns))

	for row in dataframe_to_rows(mod_output_df, index=False, header=False):
		ws.append(row)

	for cell in ws[1]:
		cell.font = global_vars.font_header

	for cell in ws[1] + ws[len(mod_output_df)+1]:
		cell.border = Border(top=global_vars.border_APA, bottom=global_vars.border_APA)
	for cell in ws[len(mod_output_df)+1]:
		cell.border = Border(bottom=global_vars.border_APA)

	for row in range(1, len(mod_output_df)+2):
		for cell in ws[row]:
			cell.alignment = global_vars.alignment_center

	table_notes = ["R squared adjusted = {R}".format(R="{:.2f}".format(output_df["R2adj"][0])), "Dependent Variable: {DV}".format(DV=global_vars.raw_mr_outcomevar)]
	helper_funcs.add_table_notes(ws, table_notes)

	#helper_funcs.save_file("raw_data_MR", wb)
	wb.save(filename=global_vars.output_filename + ".xlsx")

'''
def main():
	raw_data_df = get_raw_data_df()
	mod_raw_data_df = modify_raw_data_df(raw_data_df)
	output_df = generate_output_df(mod_raw_data_df)
	output_df = helper_funcs.multitest_correction(output_df)
	save_output(mod_raw_data_df, output_df)
'''