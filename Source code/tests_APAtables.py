# MUltiple tests corrections and FOrmatted tables Software (MUFOS)
# Copyright (C) 2020  Nikolay Petrov, Vasil Atanasov, & Trevor Thompson
import pandas as pd
import global_vars
import decision_funcs
import os
import traceback

from docx import Document
from openpyxl import load_workbook

# Suprpressing SettingWithCopyWarning pandas warning
pd.options.mode.chained_assignment = None

# All expected APA tables are located in the APA_tables/expected tables folder. The tests are located in the Apa-table_tests.csv file - this file provides input values
# for each global variable.
# The tests below go through each test from the csv, set the global variables, use the deicision funcs (which trigger the main funcs) to generate APA tables
# and compare the newly generated APA tables with the expected tables. For word, this is based on the xml, for excel the comparison is done
# by comparing each cell's properties
# Any of the expected/output tables files are modified manually in any way, the tests will fail. This is most likely due to excel picking up the current user's theme.
# For example, if on my computer I change the VALUE of an expected Excel table, an error will pop up on the border of the first cell (A1) because the border color will not match.

TESTS_APA_TABLES_INPUT_DATAFRAMES_FOLDER = os.path.join(global_vars.unit_tests_directory, "APA_tables", "input dataframes")
TESTS_APA_TABLES_OUTPUT_TABLES_FOLDER = os.path.join(global_vars.unit_tests_directory, "APA_tables", "output tables")
TEST_APA_TABLES_EXPECTED_TABLES_FOLDER = os.path.join(global_vars.unit_tests_directory, "APA_tables", "expected tables")

#-------------------------------------------------------Running the tests from the test.csv file
def apa_outputs_test():
	df = pd.read_csv(os.path.join(global_vars.unit_tests_directory, "APA_tables", "apa_table_tests.csv"), keep_default_na=False)

	def str_to_arr(x, pairttest=False):
		x = x.replace("[","")
		x = x.replace("]", "")
		x = x.replace("'", "")
		x = x.replace(" ", "")
		x = x.split(",")
		if pairttest:
			x = [x[i:i+2] for i in range(0, len(x), 2)]
		return x

	for row_ind in range(0, len(df)):
		row = df.loc[row_ind]
		
		global_vars.input_path_and_filename = os.path.join(TESTS_APA_TABLES_INPUT_DATAFRAMES_FOLDER, row[0])
		global_vars.alpha_threshold = row[1]
		global_vars.output_filename = None
		global_vars.output_filetype = row[3]
		global_vars.input_type = row[4]
		global_vars.raw_test = row[5]
		global_vars.raw_corr_type = row[6]
		global_vars.raw_corr_vars = str_to_arr(row[7])
		global_vars.raw_corr_include_CI = row[8]
		global_vars.raw_mr_outcomevar = row[9]
		global_vars.raw_mr_predictors = str_to_arr(row[10])
		global_vars.raw_indttest_groupvar = row[11]
		global_vars.raw_indttest_grouplevel1 = row[12]
		global_vars.raw_indttest_grouplevel2 = row[13]
		global_vars.raw_indttest_dv = str_to_arr(row[14])
		global_vars.raw_pairttest_var_pairs = str_to_arr(row[15], pairttest=True)
		global_vars.summ_corr_varOne = row[16]
		global_vars.summ_corr_varTwo = row[17]
		global_vars.summ_corr_coeff = row[18]
		global_vars.summ_corr_pvalues = row[19]
		global_vars.summ_indttest_var = row[20]
		global_vars.summ_indttest_meanOne = row[21]
		global_vars.summ_indttest_sdOne = row[22]
		global_vars.summ_indttest_nOne = row[23]
		global_vars.summ_indttest_meanTwo = row[24]
		global_vars.summ_indttest_sdTwo = row[25]
		global_vars.summ_indttest_nTwo = row[26]
		global_vars.summ_indttest_equal_var = row[27]
		global_vars.spss_test = row[28]
		global_vars.spss_indttest_nOne = row[29]
		global_vars.spss_indttest_nTwo = row[30]
		global_vars.spss_indttest_groupOneLabel = row[31]
		global_vars.spss_indttest_groupTwoLabel = row[32]
		global_vars.spss_pairttest_n = row[33]
		global_vars.pvalues_col = row[34]
		global_vars.effect_size_choice = row[35]
		global_vars.correction_type = row[36]
		global_vars.non_numeric_input_raise_errors = row[37]
		global_vars.corr_table_triangle = row[38]
		test_name = row[39]
		test_id = row[40]

		global_vars.output_filename = os.path.join(TESTS_APA_TABLES_OUTPUT_TABLES_FOLDER, str(test_id) + "_" + test_name)

		raw_data_df = decision_funcs.get_raw_data_df()
		mod_raw_data_df = decision_funcs.modify_raw_data_df(raw_data_df)
		output_df = decision_funcs.generate_output_df(mod_raw_data_df)
		output_df = decision_funcs.multitest_correction(output_df)
		decision_funcs.save_output(mod_raw_data_df, output_df)

		if global_vars.output_filetype == "Word":
			expected_table = Document(os.path.join(TEST_APA_TABLES_EXPECTED_TABLES_FOLDER, str(test_id) + "_" + test_name + ".docx"))
			output_table = Document(os.path.join(TESTS_APA_TABLES_OUTPUT_TABLES_FOLDER, str(test_id) + "_" + test_name + ".docx"))
			if expected_table.element.xml != output_table.element.xml:
				raise Exception("Problem when comparing word xml. Test ID: {}".format(test_id))
		elif global_vars.output_filetype == "Excel":
			expected_table = load_workbook(os.path.join(TEST_APA_TABLES_EXPECTED_TABLES_FOLDER, str(test_id) + "_" + test_name + ".xlsx"), read_only=True)
			expected_table_ws = expected_table.active
			output_table = load_workbook(os.path.join(TESTS_APA_TABLES_OUTPUT_TABLES_FOLDER, str(test_id) + "_" + test_name + ".xlsx"), read_only=True)
			output_table_ws = output_table.active

			# get all cell attributes; then loop through each cell and compare each cell's attributes
			cell_attrs = [attr for attr in dir(expected_table_ws["A1"]) if not attr.startswith("_") and attr != "parent"]
			for row in expected_table_ws.iter_rows(min_row=1, max_row=expected_table_ws.max_row, max_col=expected_table_ws.max_column):
				for expected_table_cell in row:
					if type(expected_table_cell).__name__ != "EmptyCell":
						output_table_cell = output_table_ws[expected_table_cell.coordinate]
						for attr in cell_attrs:
							if getattr(expected_table_cell, attr) != getattr(output_table_cell, attr):
								raise Exception("Problem when comparing excel cells. Test ID {id}. Problematic cell's coordinates: {c}. Problematic attribute: {a}".format(id=test_id, c=expected_table_cell.coordinate, a=attr))

def main():
	try:
		apa_outputs_test()
		print("APA tables tests passed!")
	except Exception:
		print(traceback.format_exc())
		print("Remember that if any of the expected/output tables files are modified manually in any way, the tests will fail.")