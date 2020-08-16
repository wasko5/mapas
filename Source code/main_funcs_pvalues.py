# MUltiple tests corrections and FOrmatted tables Software (MUFOS)
# Copyright (C) 2020  Nikolay Petrov, Vasil Atanasov, & Trevor Thompson
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import global_vars
import helper_funcs

#-----------------------------------------------------------Modified raw data dataframe----------------------------------------------------
def pvalues_generate_mod_raw_data_df(raw_data_df):
	mod_raw_data_df = raw_data_df.copy()

	return mod_raw_data_df

#-----------------------------------------------------------Output dataframe----------------------------------------------------
def pvalues_generate_output_df(mod_raw_data_df):
	output_df = mod_raw_data_df.copy()

	return output_df

#-----------------------------------------------------------Saving data----------------------------------------------------
def pvalues_table(mod_raw_data_df, output_df):
	wb = Workbook()
	ws = wb.active

	for row in dataframe_to_rows(output_df, index=False, header=True):
		ws.append(row)

	for row in range(1, len(output_df)+2):
		for cell in ws[row]:
			cell.alignment = global_vars.alignment_center

	for cell in ws[1]:
		cell.font = global_vars.font_bold

	for cell in ws[1] + ws[len(output_df)+1]:
		cell.border = Border(bottom=global_vars.border_APA)

	helper_funcs.add_table_notes(ws, [])

	wb.save(filename=global_vars.output_filename + ".xlsx")