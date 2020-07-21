import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import global_vars
import helper_funcs

#-----------------------------------------------------------Modified raw data dataframe----------------------------------------------------
def spss_corr_generate_mod_raw_data_df(raw_data_df):
	if not(raw_data_df.columns[0] == "Correlations" or
		raw_data_df.iloc[1, 0] == "Spearman's rho" or raw_data_df.iloc[1, 0] == "Kendall's tau_b" or raw_data_df.iloc[1, 1] == "Pearson Correlation" or
		"Correlation is significant at the" in raw_data_df.iloc[raw_data_df.index[-1], 0] or
		raw_data_df.iloc[raw_data_df.index[-1], :][1:].isna().all() or
		raw_data_df.columns[1:].str.contains("Unnamed:").all()):
		
		raise Exception("The provided SPSS correlation table does not seem to be in the accepted format. Please use the \"How to export SPSS table?\" button for guidance or refer to the documentation.")

	mod_raw_data_df = raw_data_df.copy()

	if mod_raw_data_df.iloc[1, 1] == "Pearson Correlation":
		mod_raw_data_df.iloc[0, 0], mod_raw_data_df.iloc[0, 1] = "Variable", "Statistic"
		mod_raw_data_df = mod_raw_data_df.rename(columns=mod_raw_data_df.iloc[0]).drop(mod_raw_data_df.index[0]).reset_index(drop=True)
	elif mod_raw_data_df.iloc[1, 0] == "Kendall's tau_b" or mod_raw_data_df.iloc[1, 0] == "Spearman's rho":
		mod_raw_data_df.iloc[0, 1], mod_raw_data_df.iloc[0, 2] = "Variable", "Statistic"
		mod_raw_data_df = mod_raw_data_df.rename(columns=mod_raw_data_df.iloc[0]).drop(mod_raw_data_df.index[0]).reset_index(drop=True)
		correlation_label = mod_raw_data_df.iloc[0, 0]
		mod_raw_data_df["Statistic"].replace(to_replace="Correlation Coefficient", value=correlation_label, inplace=True)
		mod_raw_data_df.drop(columns = list(mod_raw_data_df.columns)[0], inplace=True)

	return mod_raw_data_df

#-----------------------------------------------------------Output dataframe----------------------------------------------------
def spss_corr_generate_output_df(mod_raw_data_df):
	'''
	Logic: 1) from the variable column, get the indices of all rows that have a variable 2) loop through those indices and create a standalone dataframe
	for each variable 3) from that dataframe, look at the correlations 4) find the diagonal (correlation is 1) 5) once the diagnoal is found
	then go through each subsequent column and get the needed stats stuff based on the kys of the output_dict
	'''
	correlation_label = mod_raw_data_df.iloc[0, 1]
	output_dict = {"var1": [], "var2": [], correlation_label: [], "pvalues": []}
	vars_row_ind_list = list(mod_raw_data_df[mod_raw_data_df["Variable"].notnull()].index)

	for ind, ele in enumerate(vars_row_ind_list):
		if ind != len(vars_row_ind_list)-1: #check for overflow
			current_df = mod_raw_data_df.iloc[vars_row_ind_list[ind]: vars_row_ind_list[ind+1]]
			row_corrs = current_df.iloc[0, 2:]
			diagonal_flag = False
			for var, corr in row_corrs.items():
				if corr == 1:
					diagonal_flag = True
					continue
				if diagonal_flag == True:
					output_dict["var1"].append(current_df.iloc[0, 0])
					output_dict["var2"].append(var)
					output_dict[correlation_label].append(corr) #Note: corrs are not formatted here as multi test correction first is applied and is then re-formatted based on the updated pvalues
					output_dict["pvalues"].append(current_df.loc[ele+1, var])

	output_df = pd.DataFrame(output_dict)

	return output_df

#-----------------------------------------------------------4. Saving data----------------------------------------------------
#4.2.  Main function for saving data
def spss_corr_apa_table(mod_raw_data_df, output_df):
	#could very easily use the summ_corr_apa_table function here as I pass identical data - seperate is preferred, however, as might be updated/adjusted in the future
	#note that the raw corr apa table has CIs so not usable
	correlation_label = mod_raw_data_df.iloc[0, 1]
	variables_list = list(mod_raw_data_df.columns)[2:]

	wb = Workbook()
	ws = wb.active

	ws.append([""] + variables_list[1:])
	for ind,var in enumerate(variables_list[:-1]):
		ws.cell(row=ind+2, column=1).value = var

	start_col = 2
	for row in range(2, len(variables_list)+1):
		row_var = ws.cell(row=row, column=1).value
		for col in range(start_col, len(variables_list)+1):
			col_var = ws.cell(row=1, column=col).value
			#here query method is not only slower as it is much smaller dataset but also cannot refer to two different vaiables (colname and val)
			df_filter = output_df[((output_df["var1"]==row_var) & (output_df["var2"]==col_var)) | ((output_df["var1"]==col_var) & (output_df["var2"]==row_var))].iloc[0]
			r = df_filter[correlation_label]
			p = df_filter["adjusted_pvalues"]
			r = helper_funcs.correlations_format_val(r, p)
			ws.cell(row=row, column=col).value = r
		start_col += 1

	for row in range(1, len(variables_list)+1):
		for cell in ws[row]:
			cell.alignment = global_vars.alignment_center

	for cell in ws[1]:
		cell.font = global_vars.font_header

	for row in range(2, len(variables_list)+1):
		ws.cell(row=row, column=1).font = global_vars.font_header

	for cell in ws[len(variables_list)]:
		cell.border = Border(bottom=global_vars.border_APA)

	for cell in ws[1]:
		cell.border = Border(top=global_vars.border_APA, bottom=global_vars.border_APA)

	table_notes = ["**p < 0.01", "*p < {}".format(global_vars.alpha_threshold)]
	helper_funcs.add_table_notes(ws, table_notes)

	wb.save(filename=global_vars.output_filename + ".xlsx")