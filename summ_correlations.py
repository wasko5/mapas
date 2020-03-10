import pandas as pd
import numpy as np
from scipy import stats
import statsmodels.api as sm
from statsmodels.stats import multitest
import researchpy as rp
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import global_vars
import calculations_helper_functions_only as helper_funcs

#-----------------------------------------------------------0. Input----------------------------------------------------
#Commented out variables would either not be used or are considered for removal
global_vars.input_path_and_filename = "Input files\\input_summary_correlations.xlsx"
#input_fileext = ""
global_vars.alpha_threshold = 0.05
global_vars.output_filename = "Output files for testing\\summ_correlations_" #this does not end in .xlsx just for testing; in real app it should end with .xlsx (reason: function will add random numbers to prevent overwriting)

global_vars.input_type = "summ_corr"

global_vars.raw_test = ""

#raw_corr_type = ""
#raw_mr_outcomevar = ""
#raw_mr_predictors = ""
#raw_indttest_groupvar = ""
#raw_indttest_dv = ""
#raw_pairttest_var_pairs = ""

global_vars.summ_corr_varOne = "variable 1"
global_vars.summ_corr_varTwo = "variable 2"
global_vars.summ_corr_coeff = "test statistic r"
global_vars.summ_corr_pvalues = "pvalues"

#summ_indttest_var = ""
#summ_indttest_meanOne = ""
#summ_indttest_sdOne = ""
#summ_indttest_nOne = ""
#summ_indttest_meanTwo = ""
#summ_indttest_sdTwo = ""
#summ_indttest_nTwo = ""
#summ_indttest_equal_var = ""

#spss_test = ""
#spss_indttest_nOne = ""
#spss_indttest_nTwo = ""
#spss_indttest_groupOneLabel = ""
#spss_indttest_groupTwoLabel = ""
#spss_pairttest_nOne = ""
#spss_pairttest_nTwo = ""

#effect_size_choice = "" #could also be "Hedge's g", "Glass's delta" or blank/none (no idea which one it was)
global_vars.correction_type = "fdr_bh" #see global_vars.master_dict for other values

#non_numeric_input_raise_errors = True #defaults to True here
#raw_ttest_output_descriptives = ""


#-----------------------------------------------------------1. Main flow----------------------------------------------------
#Note that the execution of the main flow is at the bottom
def get_raw_data_df():
	if global_vars.input_path_and_filename.endswith(".xlsx"):
		try:
			raw_data_df = pd.read_excel(global_vars.input_path_and_filename)
		except:
			raise Exception("Oh-oh. For some reason we cannot read the provided file. Please try another file - make sure it's an excel spreadsheet.")

	#raw_data_df.to_excel("Progress dataframes\\summ_correlations\\raw_data_df.xlsx")

	return raw_data_df

def modify_raw_data_df(raw_data_df):
	numeric_cols = [global_vars.summ_corr_coeff, global_vars.summ_corr_pvalues]
	non_numeric_cols = [global_vars.summ_corr_varOne, global_vars.summ_corr_varTwo]
	helper_funcs.error_on_input(df=raw_data_df, cols=numeric_cols, input_type="numeric")
	helper_funcs.error_on_input(df=raw_data_df, cols=non_numeric_cols, input_type="nan")
	mod_raw_data_df = raw_input_generate_mod_raw_data_df(raw_data_df, numeric_cols, non_numeric_cols) #this function works but need to change name

	#mod_raw_data_df.to_excel("Progress dataframes\\summ_correlations\\mod_raw_data_df.xlsx")

	return mod_raw_data_df

def generate_output_df(mod_raw_data_df):
	output_df = summ_corr_generate_output_df(mod_raw_data_df)

	#output_df.to_excel("Progress dataframes\\summ_correlations\\output_df.xlsx")

	return output_df

def save_output(mod_raw_data_df, output_df):
	summ_corr_apa_table(mod_raw_data_df, output_df)

#-----------------------------------------------------------2. Modified raw data dataframe----------------------------------------------------
#2.1.  Main function for generating the modified raw data dataframe
def raw_input_generate_mod_raw_data_df(raw_data_df, numeric_cols, non_numeric_cols=[]):
	mod_raw_data_df = raw_data_df[numeric_cols + non_numeric_cols] #does NOT raise errors when array is blank (default arg)
	mod_raw_data_df[numeric_cols] = mod_raw_data_df[numeric_cols].apply(pd.to_numeric, errors="coerce") #non-numeric values will be np.nan
	try:
		mod_raw_data_df[non_numeric_cols] = mod_raw_data_df[non_numeric_cols].astype(str) #does NOT raise errors when array is blank (default arg)
	except:
		raise Exception("The data could not be processed. Please ensure that the non-numeric columns contain only strings.")
	#mod_raw_data_df = mod_raw_data_df.dropna().reset_index(drop=True) NEVER DOROP NA HERE - JUST COERCE TO KNOW WHERE NA VALUES ARE BUT DO NOT ALTER!!!!
	
	return mod_raw_data_df

#-----------------------------------------------------------3. Output dataframe----------------------------------------------------
#3.2.  Main function for generating the output data dataframe
def summ_corr_generate_output_df(mod_raw_data_df):
	#all checks are done so no need for modifications; function used only for infrastructural consistency
	output_df = mod_raw_data_df.copy()

	return output_df

#-----------------------------------------------------------4. Saving data----------------------------------------------------
#4.2.  Main function for saving data
def summ_corr_apa_table(mod_raw_data_df, output_df):
	variables_list = list(output_df[global_vars.summ_corr_varOne].value_counts().keys()) + [list(output_df[global_vars.summ_corr_varTwo].value_counts().keys())[0]]

	wb = Workbook()
	ws = wb.active

	ws.append([""] + variables_list[1:])
	for ind,var in enumerate(variables_list[:-1]):
		ws.cell(row=ind+2, column=1).value = var

	start_col = 2
	for row in range(2, len(variables_list)+1):
		row_var = ws.cell(row=row, column=1).value
		for col in range(start_col, len(variables_list)+1):
			col_var = ws.cell(row=1, column=col).value
			#here query method is not only slower as it is much smaller dataset but also cannot refer to two different vaiables (colname and val)
			r = output_df[((output_df[global_vars.summ_corr_varOne]==row_var) & (output_df[global_vars.summ_corr_varTwo]==col_var)) | ((output_df[global_vars.summ_corr_varOne]==col_var) & (output_df[global_vars.summ_corr_varTwo]==row_var))].iloc[0][global_vars.summ_corr_coeff]
			p = output_df[((output_df[global_vars.summ_corr_varOne]==row_var) & (output_df[global_vars.summ_corr_varTwo]==col_var)) | ((output_df[global_vars.summ_corr_varOne]==col_var) & (output_df[global_vars.summ_corr_varTwo]==row_var))].iloc[0]["adjusted_pvalues"]
			r = helper_funcs.correlations_format_val(r, p)
			ws.cell(row=row, column=col).value = r
		start_col += 1

	for row in range(1, len(variables_list)+1):
		for cell in ws[row]:
			cell.alignment = global_vars.alignment_center

	for cell in ws[1]:
		cell.font = global_vars.font_header

	for row in range(2, len(variables_list)+1):
		ws.cell(row=row, column=1).font = global_vars.font_header

	for cell in ws[len(variables_list)]:
		cell.border = Border(bottom=global_vars.border_APA)

	for cell in ws[1]:
		cell.border = Border(top=global_vars.border_APA, bottom=global_vars.border_APA)

	table_notes = ["**p < 0.01", "*p < {}".format(global_vars.alpha_threshold)]
	helper_funcs.add_table_notes(ws, table_notes)

	helper_funcs.save_file("summ_correlations", wb)


raw_data_df = get_raw_data_df()
mod_raw_data_df = modify_raw_data_df(raw_data_df)
output_df = generate_output_df(mod_raw_data_df)
output_df = helper_funcs.multitest_correction(output_df)
save_output(mod_raw_data_df, output_df)