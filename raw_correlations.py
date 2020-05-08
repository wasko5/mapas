import pandas as pd
import numpy as np
from scipy import stats
import statsmodels.api as sm
from statsmodels.stats import multitest
import researchpy as rp
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import global_vars
import calculations_helper_functions_only as helper_funcs

#-----------------------------------------------------------0. Input----------------------------------------------------
#TESTING ONLY
'''
global_vars.input_path_and_filename = "Input files\\raw data\\input_raw_correlations.xlsx"
global_vars.alpha_threshold = 0.05
global_vars.output_filename = "Test outputs\\output_raw_correlations" #this does not end in .xlsx just for testing; in real app it should end with .xlsx (reason: function will add random numbers to prevent overwriting)

global_vars.input_type = "raw"

global_vars.raw_test = "corr"

global_vars.raw_corr_type = "pearson" #could be "spearman" or "kendall"
global_vars.raw_corr_vars = ["var1", "var2", "var3", "var4", "var5", "var6", "var7", "var8"] #ASAP - to implement & test
#raw_mr_outcomevar = ""
#raw_indttest_groupvar = ""
#raw_pairttest_var1 = ""
#raw_pairttest_var2 = ""

#summ_corr_varOne = ""
#summ_corr_varTwo = ""
#summ_corr_coeff = ""
#summ_corr_pvalues = ""

#summ_indttest_var = ""
#summ_indttest_meanOne = ""
#summ_indttest_sdOne = ""
#summ_indttest_nOne = ""
#summ_indttest_meanTwo = ""
#summ_indttest_sdTwo = ""
#summ_indttest_nTwo = ""
#summ_indttest_equal_var = ""

#spss_test = ""
#spss_indttest_nOne = ""
#spss_indttest_nTwo = ""
#spss_indttest_groupOneLabel = ""
#spss_indttest_groupTwoLabel = ""
#spss_pairttest_nOne = ""
#spss_pairttest_nTwo = ""

#effect_size_choice = ""
global_vars.correction_type = "fdr_bh" #see global_vars.master_dict for other values

global_vars.non_numeric_input_raise_errors = True #or False
#raw_ttest_output_descriptives = ""
'''

#-----------------------------------------------------------1. Main flow----------------------------------------------------
#Note that the execution of the main flow is at the bottom
'''
#execution now guided by the decision_funcs
def get_raw_data_df():
	if global_vars.input_path_and_filename.endswith(".xlsx"):
		try:
			raw_data_df = pd.read_excel(global_vars.input_path_and_filename)
		except:
			raise Exception("Oh-oh. For some reason we cannot read the provided file. Please try another file - make sure it's an excel spreadsheet.")

	#mod_raw_data_df.to_excel("Progress dataframes\\raw_correlations\\raw_data_df.xlsx")

	return raw_data_df

def modify_raw_data_df(raw_data_df):
	numeric_cols = global_vars.raw_corr_vars
	if global_vars.non_numeric_input_raise_errors == True:
		helper_funcs.error_on_input(df=raw_data_df, cols=numeric_cols, input_type="numeric")
	
	mod_raw_data_df = raw_input_generate_mod_raw_data_df(raw_data_df, numeric_cols)

	#mod_raw_data_df.to_excel("Progress dataframes\\raw_correlations\\mod_raw_data_df.xlsx")

	return mod_raw_data_df

def generate_output_df(mod_raw_data_df):
	output_df = raw_corr_generate_output_df(mod_raw_data_df)

	#output_df.to_excel("Progress dataframes\\raw_correlations\\output_df.xlsx")

	return output_df

def save_output(mod_raw_data_df, output_df):
	raw_corr_apa_table(mod_raw_data_df, output_df)
'''
#-----------------------------------------------------------2. Modified raw data dataframe----------------------------------------------------
#2.1.  Main function for generating the modified raw data dataframe
'''
#this is now done in the helper_funcs
def raw_input_generate_mod_raw_data_df(raw_data_df, numeric_cols, non_numeric_cols=[]):
	mod_raw_data_df = raw_data_df[numeric_cols + non_numeric_cols] #does NOT raise errors when array is blank (default arg)
	mod_raw_data_df[numeric_cols] = mod_raw_data_df[numeric_cols].apply(pd.to_numeric, errors="coerce") #non-numeric values will be np.nan
	try:
		mod_raw_data_df[non_numeric_cols] = mod_raw_data_df[non_numeric_cols].astype(str) #does NOT raise errors when array is blank (default arg)
	except:
		raise Exception("The data could not be processed. Please ensure that the non-numeric columns contain only strings.")
	#mod_raw_data_df = mod_raw_data_df.dropna().reset_index(drop=True) NEVER DOROP NA HERE - JUST COERCE TO KNOW WHERE NA VALUES ARE BUT DO NOT ALTER!!!!
	
	return mod_raw_data_df
'''
#-----------------------------------------------------------3. Output dataframe----------------------------------------------------
#3.2.  Main function for generating the output  data dataframe
def raw_corr_generate_output_df(mod_raw_data_df):
	'''
	The critical moment here is that I only have np.nan values where values are non-numeric - can't have anything different from np.nan or a numeric value - bad things happen.
	It's also important to use this type of loop otherwise there is inconsistent behaviour when dropping nan values:
		#https://github.com/scipy/scipy/issues/9103
	Might be worth investigating creating a custom dropping algorithm here as the in-built ones seem to be slow
		#https://github.com/scipy/scipy/issues/6654
	'''
	data_list = []

	i=0
	for var1 in mod_raw_data_df.columns: #can use .columns as all columns are now only of relevant vars (done in raw_input_generate_mod_raw_data_df func)
		for var2 in mod_raw_data_df.columns[i:]:
			if not var1 == var2:
				r, p = helper_funcs.raw_input_corr_coeff(mod_raw_data_df[var1],mod_raw_data_df[var2])
				ci_low, ci_high = helper_funcs.corr_coeff_ci(r, n=min(mod_raw_data_df[var1].count(), mod_raw_data_df[var2].count())) #takes the smallest number as this would be the number of correlations done
				data_list.append((var1, var2, r, ci_low, ci_high, p))
		i += 1
	output_df = pd.DataFrame(data_list, columns=["Variable1", "Variable2", "Correlation_Coefficient", "CI_low", "CI_high", "pvalues"])

	return output_df

#-----------------------------------------------------------4. Saving data----------------------------------------------------
#4.2.  Main function for saving data
def raw_corr_apa_table(mod_raw_data_df, output_df):
	wb = Workbook()
	ws = wb.active
	
	variables_list = list(mod_raw_data_df.columns) #local var and .columns preferred to minimize use of global scope
	
	ws.append([""] + variables_list[:-1] + ["Mean", "SD"])
	ws["A2"] = variables_list[0]
	for ind,var in enumerate(variables_list[1:]):
		ws.cell(row=(ind*2)+3, column=1).value = var
	
	start_row = 3
	for col in range(2, len(variables_list)+3):
		col_var = ws.cell(row=1, column=col).value
		for row in range(start_row, len(variables_list)*2, 2):
			row_var = ws.cell(row=row, column=1).value
			#Using the query method is slower than using conditionals - in my tests query was between 65 and 80 % slower for 1,000 and 10,000 runs
			r = output_df[((output_df["Variable1"]==row_var) & (output_df["Variable2"]==col_var)) | ((output_df["Variable1"]==col_var) & (output_df["Variable2"]==row_var))].iloc[0]["Correlation_Coefficient"]
			p = output_df[((output_df["Variable1"]==row_var) & (output_df["Variable2"]==col_var)) | ((output_df["Variable1"]==col_var) & (output_df["Variable2"]==row_var))].iloc[0]["adjusted_pvalues"]
			ci_low = output_df[((output_df["Variable1"]==row_var) & (output_df["Variable2"]==col_var)) | ((output_df["Variable1"]==col_var) & (output_df["Variable2"]==row_var))].iloc[0]["CI_low"]
			ci_high = output_df[((output_df["Variable1"]==row_var) & (output_df["Variable2"]==col_var)) | ((output_df["Variable1"]==col_var) & (output_df["Variable2"]==row_var))].iloc[0]["CI_high"]
			r = helper_funcs.correlations_format_val(r, p)
			ci_low, ci_high = helper_funcs.correlations_format_val(ci_low), helper_funcs.correlations_format_val(ci_high)
			ws.cell(row=row, column=col).value = r
			ws.cell(row=row+1, column=col).value = "[" + ci_low + ", " + ci_high + "]"
		start_row += 2

	ws.cell(row=2, column=len(variables_list)+1).value = "{:.2f}".format(mod_raw_data_df[variables_list[0]].mean(skipna=True))
	ws.cell(row=2, column=len(variables_list)+2).value = "{:.2f}".format(mod_raw_data_df[variables_list[0]].std(skipna=True))
	for row in range(3, len(variables_list)*2, 2):
		ws.cell(row=row, column=len(variables_list)+1).value = "{:.2f}".format(mod_raw_data_df[variables_list[int((row-1)/2)]].mean(skipna=True))
		ws.cell(row=row, column=len(variables_list)+2).value = "{:.2f}".format(mod_raw_data_df[variables_list[int((row-1)/2)]].std(skipna=True))    
	
	for row in range(3, len(variables_list)*2, 2):
		ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=1)
		ws.merge_cells(start_row=row, start_column=len(variables_list)+1, end_row=row+1, end_column=len(variables_list)+1)
		ws.merge_cells(start_row=row, start_column=len(variables_list)+2, end_row=row+1, end_column=len(variables_list)+2)
		
	for row in ws[1:len(variables_list)*2]:
		for cell in row:
			if cell.row > 2 and (cell.column == 1 or cell.column == len(variables_list)+1 or cell.column == len(variables_list)+2):
				cell.alignment = global_vars.alignment_top
			else:
				cell.alignment = global_vars.alignment_center
				
	for cell in ws[1]:
		cell.font = global_vars.font_header
	
	ci_font = Font(size=9)
	for row in range(4, len(variables_list)*2+1, 2):
		for cell in ws[row]:
			cell.font = ci_font
	
	for cell in ws[1]:
		cell.border = Border(top=global_vars.border_APA, bottom=global_vars.border_APA)
	for cell in ws[len(variables_list)*2]:
		cell.border = Border(bottom = global_vars.border_APA)
	
	table_notes = ["Note. M and SD represent mean and standard deviation, respectively. Values in square brackets indicate 95% confidence interval for the correlation coefficient."]
	#might change the below to a separate lookup func
	table_notes.append("Correlation coefficient used: {}".format(list(global_vars.master_dict.keys())[list(global_vars.master_dict.values()).index(global_vars.raw_corr_type)]))
	table_notes.append("**p < 0.01")
	table_notes.append("*p < {}".format(global_vars.alpha_threshold))
	helper_funcs.add_table_notes(ws, table_notes) 
	
	#helper_funcs.save_file("raw_data_correlations", wb)
	wb.save(filename=global_vars.output_filename + ".xlsx")

'''
def main():
	raw_data_df = get_raw_data_df()
	mod_raw_data_df = modify_raw_data_df(raw_data_df)
	output_df = generate_output_df(mod_raw_data_df)
	output_df = helper_funcs.multitest_correction(output_df)
	save_output(mod_raw_data_df, output_df)
'''