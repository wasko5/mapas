import pandas as pd
import numpy as np
from scipy import stats
import statsmodels.api as sm
from statsmodels.stats import multitest
import researchpy as rp
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import global_vars
import calculations_helper_functions_only as helper_funcs

#-----------------------------------------------------------0. Input----------------------------------------------------
#TESTING ONLY
'''
#global_vars.input_path_and_filename = "Input files\\spss\\input_spss_pairedTtest.xlsx"
global_vars.input_path_and_filename = "C:\\Users\\Nikolay Petrov\\Desktop\\testpairttest.xlsx"
global_vars.alpha_threshold = 0.05
global_vars.output_filename = "Output files for testing\\spss_pairttest" #this does not end in .xlsx just for testing; in real app it should end with .xlsx (reason: function will add random numbers to prevent overwriting)

global_vars.input_type = "spss"

global_vars.raw_test = ""

#raw_corr_type = ""
#raw_mr_outcomevar = ""
#raw_mr_predictors = ""
#raw_indttest_groupvar = ""
#raw_indttest_dv = ""
#raw_pairttest_var_pairs = ""

#summ_corr_varOne = ""
#summ_corr_varTwo = ""
#summ_corr_coeff = ""
#summ_corr_pvalues = ""

#summ_indttest_var = ""
#summ_indttest_meanOne = ""
#summ_indttest_sdOne = ""
#summ_indttest_nOne = ""
#summ_indttest_meanTwo = ""
#summ_indttest_sdTwo = ""
#summ_indttest_nTwo = ""
#summ_indttest_equal_var = ""

global_vars.spss_test = "pairttest"
#spss_indttest_nOne = ""
#spss_indttest_nTwo = ""
#spss_indttest_groupOneLabel = ""
#spss_indttest_groupTwoLabel = ""
global_vars.spss_pairttest_nOne = 135
global_vars.spss_pairttest_nTwo = 125

global_vars.effect_size_choice = "Cohen's d"
global_vars.correction_type = "bonferroni" #see global_vars.master_dict for other values

#global_vars.non_numeric_input_raise_errors = True #or False
#raw_ttest_output_descriptives = ""
'''

#-----------------------------------------------------------1. Main flow----------------------------------------------------
#Note that the execution of the main flow is at the bottom
'''
def get_raw_data_df():
	if global_vars.input_path_and_filename.endswith(".xlsx"):
		try:
			raw_data_df = pd.read_excel(global_vars.input_path_and_filename)
		except:
			raise Exception("Oh-oh. For some reason we cannot read the provided file. Please try another file - make sure it's an excel spreadsheet.")

	#raw_data_df.to_excel("Progress dataframes\\spss_pairttest\\raw_data_df.xlsx")

	return raw_data_df

def modify_raw_data_df(raw_data_df):
	mod_raw_data_df = spss_pairttest_generate_mod_raw_data_df(raw_data_df) #might remove the function depending on whether I add logical checks

	#mod_raw_data_df.to_excel("Progress dataframes\\spss_pairttest\\mod_raw_data_df.xlsx")

	return mod_raw_data_df

def generate_output_df(mod_raw_data_df):
	output_df = spss_pairttest_generate_output_df(mod_raw_data_df)

	#output_df.to_excel("Progress dataframes\\spss_pairttest\\output_df.xlsx")

	return output_df

def save_output(mod_raw_data_df, output_df):
	spss_pairttest_apa_table(mod_raw_data_df, output_df)
'''

#-----------------------------------------------------------2. Modified raw data dataframe----------------------------------------------------
#2.1.  Main function for generating the modified raw data dataframe
def spss_pairttest_generate_mod_raw_data_df(raw_data_df):
	if (len(raw_data_df.columns) != 10 or raw_data_df.iloc[1,2] != "Mean" or raw_data_df.iloc[1,3] != "Std. Deviation"
			or raw_data_df.iloc[1,4] != "Std. Error Mean" or raw_data_df.iloc[1,5] != "95% Confidence Interval of the Difference"
			or raw_data_df.iloc[2,6] != "Upper" or raw_data_df.iloc[0,7] != "t" or raw_data_df.iloc[0,8] != "df"
			or raw_data_df.iloc[0,9] != "Sig. (2-tailed)" or len(raw_data_df) <= 3):
		raise Exception("The data provided could not be read correctly. Please see the documentation for help.")
		
	mod_raw_data_df = raw_data_df.copy()
	renaming_dict = {}
	new_col_names = [mod_raw_data_df.columns[0]] + ["Variables", "Mean difference", "Std difference", "Std error difference"] + [mod_raw_data_df.iloc[1, 5]] + ["Upper", "t", "df", "p"]
	for i in range(0, len(mod_raw_data_df.columns)):
		renaming_dict[mod_raw_data_df.columns[i]] = new_col_names[i]

	mod_raw_data_df = mod_raw_data_df.rename(columns=renaming_dict).drop([mod_raw_data_df.index[0], mod_raw_data_df.index[1], mod_raw_data_df.index[2]]).reset_index(drop=True)

	return mod_raw_data_df

#-----------------------------------------------------------3. Output dataframe----------------------------------------------------
#3.2.  Main function for generating the output data dataframe
def spss_pairttest_generate_output_df(mod_raw_data_df):
	output_dict={}
	t_stat_list = list(mod_raw_data_df["t"])
	effectsize_list = []

	if global_vars.effect_size_choice == "Cohen's d":
		for t in t_stat_list:
			effect_size = t*np.sqrt(1/global_vars.spss_pairttest_nOne + 1/global_vars.spss_pairttest_nTwo)
			effectsize_list.append(effect_size)
	elif global_vars.effect_size_choice == "Hedge's g":
		#calculated using cohens_d * (1 - (3/4(n1+n2-9))) where cohens d is t*sqrt(1/n1 + 1/n2)
		for t in t_stat_list:
			effect_size = (t*np.sqrt(1/global_vars.spss_pairttest_nOne + 1/global_vars.spss_pairttest_nTwo))*np.sqrt(1/global_vars.spss_pairttest_nOne + 1/global_vars.spss_pairttest_nTwo)
			effectsize_list.append(effect_size)
	elif global_vars.effect_size_choice == "None":
		effectsize_list = [np.nan] * len(t_stat_list)
			
	output_dict["Variables"] = list(mod_raw_data_df["Variables"])
	output_dict["Time 1_Mean"] = [""] * len(output_dict["Variables"])
	output_dict["Time 1_SD"] = [""] * len(output_dict["Variables"])
	output_dict["Time 2_Mean"] = [""] * len(output_dict["Variables"])
	output_dict["Time 2_SD"] = [""] * len(output_dict["Variables"])
	output_dict["df"] = list(mod_raw_data_df["df"])
	output_dict["t"] = t_stat_list
	output_dict[global_vars.effect_size_choice] = effectsize_list
	output_dict["pvalues"] = list(mod_raw_data_df["p"])

	output_df = pd.DataFrame(output_dict)
	
	return output_df

#-----------------------------------------------------------4. Saving data----------------------------------------------------
#4.2.  Main function for saving data
def spss_pairttest_apa_table(mod_raw_data_df, output_df):
	output_df.drop(columns=["pvalues", list(output_df.columns)[-1]], inplace=True)

	pd.options.mode.chained_assignment = None
	output_df[list(output_df.columns)[5:-1]] = output_df[list(output_df.columns)[5:-1]].applymap(lambda x: "{:.2f}".format(x))

	output_df["adjusted_pvalues"] = output_df["adjusted_pvalues"].map(helper_funcs.pvalue_formatting)
	pd.options.mode.chained_assignment = "warn"

	wb = Workbook()
	ws = wb.active
	
	ws.cell(row=1, column=1).value = "Variable"
	ws.merge_cells('A1:A2')
	ws.cell(row=1, column=1).font = global_vars.font_header
	
	ws.cell(row=1, column=2).value = "Time 1"
	ws.merge_cells('B1:C1')
	
	ws.cell(row=1, column=4).value = "Time 2"
	ws.merge_cells('D1:E1')
	
	ws.cell(row=1, column=6).value = "df"
	ws.merge_cells('F1:F2')
	ws.cell(row=1, column=6).font = global_vars.font_header
	
	ws.cell(row=1, column=7).value = "t"
	ws.merge_cells('G1:G2')
	ws.cell(row=1, column=7).font = global_vars.font_header
	
	ws.cell(row=1, column=8).value = global_vars.effect_size_choice
	ws.merge_cells('H1:H2')
	ws.cell(row=1, column=8).font = global_vars.font_header
	
	ws.cell(row=1, column=9).value = "p"
	ws.merge_cells('I1:I2')
	ws.cell(row=1, column=9).font = global_vars.font_header
	
	for col in range(2,5,2):
		ws.cell(row=2, column=col).value = "M"
		ws.cell(row=2, column=col).font = global_vars.font_header
		ws.cell(row=2, column=col+1).value = "SD"
		ws.cell(row=2, column=col+1).font = global_vars.font_header
	
	for row in dataframe_to_rows(output_df, index=False, header=False):
		ws.append(row)

	for row in range(1, len(output_df)+3):
		for cell in ws[row]:
			cell.alignment = global_vars.alignment_center
	
	for cell in ws[2] + ws[len(output_df)+2]:
		cell.border = Border(bottom=global_vars.border_APA)

	table_notes = ["Means and Standard Deviations cannot be read from the SPSS table. Please add them yourself or remove those columns if they are not needed."]
	helper_funcs.add_table_notes(ws, table_notes)

	#helper_funcs.save_file("spss_data_pairttest", wb)
	wb.save(filename=global_vars.output_filename + ".xlsx")
'''
def main():
	raw_data_df = get_raw_data_df()
	mod_raw_data_df = modify_raw_data_df(raw_data_df)
	output_df = generate_output_df(mod_raw_data_df)
	output_df = helper_funcs.multitest_correction(output_df)
	save_output(mod_raw_data_df, output_df)
'''