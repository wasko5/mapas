import pandas as pd
import numpy as np
from scipy import stats
import statsmodels.api as sm
from statsmodels.stats import multitest
import researchpy as rp
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


#------------------------------------------------------------0. Additional helper functions-----------------------------------------
def get_df_columns(path_and_filename):
	try:
		if path_and_filename.endswith(".xlsx"):
			df_columns = list(pd.read_excel(path_and_filename, sheet_name=0).columns)
		elif path_and_filename.endsiwth(".csv"):
			df_columns = list(pd.read_csv(path_and_filename).columns)

		return df_columns
	except:
		raise Exception("There was a problem reading the columns from the file.")


#-----------------------------------------------------------1. Data transformations----------------------------------------------------
#-----------------1.1. Individual functions for generating modified raw data dataframes
#1.1.1. Helper functions fo generating modified raw data dataframes
def get_numeric_cols(raw_data_df):
	numeric_cols = list(raw_data_df.columns)
	if raw_test == "indttest":
		try:
			numeric_cols.remove(raw_indttest_groupvar)
		except ValueError:
			raise Exception("The grouping variable \'{}\' is not found in the file. Please make sure it is type correctly.".format(raw_indttest_groupvar))
	elif input_type == "summ_corr":
		#no need for error handling as the summ_corr_colNames_check func has already taken care of that
		numeric_cols.remove(summ_corr_varOne)
		numeric_cols.remove(summ_corr_varTwo)
	elif input_type == "summ_indttest":
		numeric_cols.remove(summ_indttest_var)
		numeric_cols.remove(summ_indttest_equal_var)
		
	return numeric_cols

def error_on_non_numeric_input(raw_data_df, numeric_cols):
	bad_vals_dict = {}
	for var in numeric_cols:
		ind_list = list(pd.to_numeric(raw_data_df[var], errors='coerce').isnull().to_numpy().nonzero()[0])
		if ind_list != []:
			bad_vals_dict[var] = ind_list
	if bad_vals_dict != {}:
		error_msg = "Non-numerical or blank entries found in the data!\n"
		for col, ind_arr in bad_vals_dict.items():
			error_msg += "In column {c}, there are non-numerical/blank entries on the following rows: {i}\n".format(c=col, i=(", ").join([str(x+2) for x in ind_arr]))
		raise Exception(error_msg)

def raw_input_pairttest_colNames_check(raw_data_df):
	inputVars_unique = []
	for var_time1, var_time2 in raw_pairttest_inputVars_list:
		if var_time1 not in raw_data_df.columns:
			raise Exception("The entered column \'{}\' is not found in the provided file.".format(var_time1))
		elif var_time2 not in raw_data_df.columns:
			raise Exception("The entered column \'{}\' is not found in the provided file.".format(var_time2))
		else:
			inputVars_unique.append(var_time1)
			inputVars_unique.append(var_time2)
	inputVars_unique = list(set(inputVars_unique))
	if len(inputVars_unique) > len(list(raw_data_df.columns)):
		raise Exception("The provided number of unique variable columns is {i}, while the number of columns in the provided file is {f}".format(i=len(inputVars_unique), f=len(list(raw_data_df.columns))))

def summ_input_colNames_check(raw_data_df, provided_cols):
	for var in provided_cols:
		if var not in list(raw_data_df.columns):
			raise Exception("Column {} is not found in the provided file.\nPlease ensure that the column names are typed correctly.".format(var))
	if len(list(raw_data_df.columns)) > len(provided_cols):
		raise Exception("The provided file contains too many columns. Please provide only {}.".format(len(provided_cols)))

#1.1.2. Individual functions for generating modified raw data dataframes
def raw_input_generate_mod_raw_data_df(raw_data_df, numeric_cols):
	if raw_test == "pairttest":
		#Treaing paired ttest input separately (as separate dataframes based on pairs) as it's possible for 1 pair 
		#to have 200 entires and another one with 400 entries
		df_list = []
		for var_time1, var_time2 in raw_pairttest_inputVars_list:
			df = raw_data_df[[var_time1, var_time2]]
			df = df.apply(pd.to_numeric, errors="coerce").dropna().reset_index(drop=True)
			df_list.append(df)
		mod_raw_data_df = pd.concat(df_list, axis=1)
	else:  
		mod_raw_data_df = raw_data_df.copy()
		mod_raw_data_df[numeric_cols] = mod_raw_data_df[numeric_cols].apply(pd.to_numeric, errors="coerce")
		mod_raw_data_df = mod_raw_data_df.dropna().reset_index(drop=True)
	
	return mod_raw_data_df

def summ_input_generate_mod_raw_data_df(raw_data_df):
	non_numeric_cols = [summ_corr_varOne, summ_corr_varTwo] if input_type == "summ_corr" else [summ_indttest_var, summ_indttest_equal_var]
	for col in non_numeric_cols:
		if True in raw_data_df[col].isnull().values:
			raise Exception("Blank row found in column \'{}\'.\nPleasu ensure there are no blank datapoints in the provided file.".format(col))
	if input_type == "summ_indttest":
		if not raw_data_df[summ_indttest_equal_var].isin([True, False]).all():
			raise Exception("A value different from True or False is found in column \'{}\'.\nPlease ensure that only True and False values are provided.".format(summ_indttest_equal_var))
	
	mod_raw_data_df = raw_data_df.copy()

	return mod_raw_data_df

def spss_corr_generate_mod_raw_data_df(raw_data_df):
	mod_raw_data_df = raw_data_df.copy()
	mod_raw_data_df.iloc[0][0], mod_raw_data_df.iloc[0][1] = "Variable", "Statistic"
	mod_raw_data_df = mod_raw_data_df.rename(columns=mod_raw_data_df.iloc[0]).drop(mod_raw_data_df.index[0]).reset_index(drop=True)
	mod_raw_data_df = mod_raw_data_df[mod_raw_data_df[mod_raw_data_df.columns[1]] != "N"].reset_index(drop=True)

	return mod_raw_data_df

def spss_mr_generate_mod_raw_data_df(raw_data_df):
	mod_raw_data_df = raw_data_df.copy()

	#renaming columns as the multi-level formatting of the spss table makes it difficult to work with
	renaming_dict = {}
	for ind, col_name in enumerate(mod_raw_data_df.columns):
		if "Unnamed" in col_name:
			if not pd.isna(mod_raw_data_df.iloc[0, ind]):
				renaming_dict[col_name] = mod_raw_data_df.iloc[0, ind]
			elif not pd.isna(mod_raw_data_df.iloc[1, ind]):
				renaming_dict[col_name] = mod_raw_data_df.iloc[1, ind]

	mod_raw_data_df = mod_raw_data_df.rename(columns=renaming_dict).drop([mod_raw_data_df.index[0], mod_raw_data_df.index[1]]).reset_index(drop=True)

	if mod_raw_data_df[mod_raw_data_df.columns[1]][0] != "(Constant)":
		raise Exception("There was a problem reading the variable column. Please try entering a correct data file. Please see the documentation for help.")
	fields_to_check = ["Unstandardized Coefficients", "Standardized Coefficients", "t", "Sig."]
	for field in fields_to_check:
		if field not in mod_raw_data_df.columns:
			raise Exception("The column \'{}\' was not found in the file. Please try entering a correct data file. Please see the documentation for help.".format(field))

	return mod_raw_data_df

def spss_indttest_generate_mod_raw_data_df(raw_data_df):
	mod_raw_data_df = raw_data_df.copy()

	#renaming columns as the multi-level formatting of the spss table makes it difficult to work with
	renaming_dict = {}
	new_col_names = list(mod_raw_data_df.columns[:2]) + list(mod_raw_data_df.iloc[1][2:-1]) + [mod_raw_data_df.columns[-1]]
	for i in range(0, len(mod_raw_data_df.columns)):
		renaming_dict[mod_raw_data_df.columns[i]] = new_col_names[i]

	mod_raw_data_df = mod_raw_data_df.rename(columns=renaming_dict).drop([mod_raw_data_df.index[0], mod_raw_data_df.index[1]]).reset_index(drop=True)
	
	if mod_raw_data_df.columns[0] != "Independent Samples Test":
		raise Exception("The SPSS table provided does not seem to be one from Independent Samples t-test. Please try with another file or refer to the documentation for help.")
	#if "Confidence Interval of the Difference" not in mod_raw_data_df.columns[9]:
	#	raise Exception("Confidence Intervals could not be read from the table. Please try with another file or refer to the instructions...")
	fields_to_check = ["F", "Sig.", "t", "df", "Sig. (2-tailed)"]
	for field in fields_to_check:
		if field not in mod_raw_data_df.columns:
			raise Exception("The column \'{}\' was not found in the file. Please try entering a correct data file. Please see the documentation for help.".format(field))

	return mod_raw_data_df

def spss_pairttest_generate_mod_raw_data_df(raw_data_df):
	if (len(raw_data_df.columns) != 10 or raw_data_df.iloc[1,2] != "Mean" or raw_data_df.iloc[1,3] != "Std. Deviation"
			or raw_data_df.iloc[1,4] != "Std. Error Mean" or raw_data_df.iloc[1,5] != "95% Confidence Interval of the Difference"
			or raw_data_df.iloc[2,6] != "Upper" or raw_data_df.iloc[0,7] != "t" or raw_data_df.iloc[0,8] != "df"
			or raw_data_df.iloc[0,9] != "Sig. (2-tailed)" or len(raw_data_df) <= 3):
		raise Exception("The data provided could not be read correctly. Please see the documentation for help.")
		
	mod_raw_data_df = raw_data_df.copy()
	renaming_dict = {}
	new_col_names = [mod_raw_data_df.columns[0]] + ["Variables", "Mean difference", "Std difference", "Std error difference"] + [mod_raw_data_df.iloc[1, 5]] + ["Upper", "t", "df", "p"]
	for i in range(0, len(mod_raw_data_df.columns)):
		renaming_dict[mod_raw_data_df.columns[i]] = new_col_names[i]

	mod_raw_data_df = mod_raw_data_df.rename(columns=renaming_dict).drop([mod_raw_data_df.index[0], mod_raw_data_df.index[1], mod_raw_data_df.index[2]]).reset_index(drop=True)

	return mod_raw_data_df

#-----------------1.2. Individual functions for generating output dataframes
#1.2.1. Helper functions for generating output dataframes
def raw_input_corr_coeff(x, y):
	if raw_corr_type == "pearson":
		x = np.array(x)
		y = np.array(y)
		nas = ~np.logical_or(np.isnan(x), np.isnan(y))
		x = np.compress(nas, x)
		y = np.compress(nas, y)

		r, p = stats.pearsonr(x, y)
	elif raw_corr_type == "spearman":
		r, p = stats.spearmanr(x, y, nan_policy="omit")
	elif raw_corr_type == "kendall":
		r, p = stats.kendalltau(x, y, nan_policy="omit")

	return r, p

def corr_coeff_ci(r, n):
	if n < 4:
		raise Exception("Sample size is too low for correlations. Minimum of 4 observations per variable required.")

	r_z = np.arctanh(r)
	se = 1/np.sqrt(n-3)
	z = stats.norm.ppf(1-0.05/2)
	low_z = r_z - z*se 
	high_z = r_z + z*se
	low, high = np.tanh((low_z, high_z))
	
	return low, high

#1.2.2. Individual functions for generating output dataframes
def raw_corr_generate_output_df(mod_raw_data_df):
	data_list = []

	i=0
	for var1 in mod_raw_data_df.columns:
		for var2 in mod_raw_data_df.columns[i:]:
			if not var1 == var2:
				r, p = raw_input_corr_coeff(mod_raw_data_df[var1],mod_raw_data_df[var2])
				ci_low, ci_high = corr_coeff_ci(r, n=mod_raw_data_df[var1].count())
				data_list.append((var1, var2, r, ci_low, ci_high, p))
		i += 1
	output_df = pd.DataFrame(data_list, columns=["Variable1", "Variable2", "Correlation Coefficient", "CI_low", "CI_high", "pvalues"])

	return output_df

def raw_mr_generate_output_df(mod_raw_data_df):
	independent_vars = list(mod_raw_data_df.columns)
	try:
		independent_vars.remove(raw_mr_outcomevar)
	except ValueError:
		raise Exception("The dependent variable \'{}\' is not found in the file. Please make sure it is typed correctly.".format(raw_mr_outcomevar))
	
	formula = raw_mr_outcomevar + "~"
	for var in independent_vars:
		formula = formula + var + "+"
	formula = formula[:-1] #the loop leaves a trailing "+" sign so this gets rid of it

	result = sm.OLS.from_formula(formula=formula, data=mod_raw_data_df).fit() #everything before the .fit() is model specification
	
	#standardization of the raw data dataframe so it can also produce beta values for the output
	mod_raw_data_df_z = mod_raw_data_df.apply(stats.zscore)
	result_z = sm.OLS.from_formula(formula=formula, data=mod_raw_data_df_z).fit()
	beta_list = list(result_z.params)

	CI_B_list = []
	for index, row in result.conf_int().iterrows(): #weirdly enough result_conf_int() returns a dataframe with the CIs
		CI_B_list.append(list(row))
		
	CI_beta_list = []
	for index, row in result_z.conf_int().iterrows():
		CI_beta_list.append(list(row))
	
	output_df = pd.DataFrame()
	output_df["Variable"] = ["(Constant)"] + independent_vars
	output_df["B"] = ["{:.2f}".format(val) for val in list(result.params)]
	output_df["Std Err B"] = ["{:.2f}".format(val) for val in list(result.bse)]
	output_df["95% CI B"] = ["[" + ",".join("{:.2f}".format(e) for e in pair) + "]" for pair in CI_B_list]
	output_df["beta"] = ["{:.2f}".format(val) for val in beta_list]
	output_df["Std Err beta"] = ["{:.2f}".format(val) for val in list(result_z.bse)] #might not need - if so, amend the apa_table func
	output_df["95% CI beta"] = ["[" + ",".join("{:.2f}".format(e) for e in pair) + "]" for pair in CI_beta_list] #might not need
	output_df["t"] = ["{:.2f}".format(val) for val in list(result.tvalues)]
	output_df["pvalues"] = list(result.pvalues)
	output_df["R2"] = ["{:.2f}".format(result.rsquared)] + [""] * (len(output_df)-1)
	output_df["R2adj"] = ["{:.2f}".format(result.rsquared_adj)] + [""] * (len(output_df)-1)
	output_df["Dependent Variable"] = [raw_mr_outcomevar] + [""] * (len(output_df)-1)
	
	return output_df

def raw_indttest_generate_output_df(mod_raw_data_df):
	unique_groups = mod_raw_data_df[raw_indttest_groupvar].unique()
	if len(unique_groups) > 2:
		raise Exception("More than two unique values found in the grouping variable \'{}\'. Please supply data with only two groups.".format(raw_indttest_groupvar))
	group1_label = str(unique_groups[0])
	group2_label = str(unique_groups[1])

	group1_df = mod_raw_data_df[mod_raw_data_df[raw_indttest_groupvar].astype(str)==group1_label]
	group2_df = mod_raw_data_df[mod_raw_data_df[raw_indttest_groupvar].astype(str)==group2_label]

	variables_list = list(mod_raw_data_df.columns)
	variables_list.remove(raw_indttest_groupvar)

	all_n_list = []
	all_mean_list = []
	all_sd_list = []
	all_se_list = []
	all_ci_low_list = []
	all_ci_high_list = []
	group1_n_list = []
	group1_mean_list = []
	group1_sd_list = []
	group1_se_list = []
	group1_ci_low_list = []
	group1_ci_high_list = []
	group2_n_list = []
	group2_mean_list = []
	group2_sd_list = []
	group2_se_list = []
	group2_ci_low_list = []
	group2_ci_high_list = []
	equal_var_list = []
	deg_free_list = []
	t_stat_list = []
	effectsize_list = []
	pvalues_list = []
	
	if effect_size_choice == "Cohen's d":
		effect_size_dfrow = 6
	elif effect_size_choice == "Hedge's g":
		effect_size_dfrow = 7
	elif effect_size_choice == "Glass's delta":
		effect_size_dfrow = 8

	for var in variables_list:
		ttest_stats_df1 = rp.ttest(group1_df[var], group2_df[var], group1_name=group1_label, group2_name=group2_label, 
			 equal_variances=stats.levene(group1_df[var], group2_df[var])[1]>0.05, paired=False)[0]
		ttest_stats_df2 = rp.ttest(group1_df[var], group2_df[var], group1_name=group1_label, group2_name=group2_label, 
			 equal_variances=stats.levene(group1_df[var], group2_df[var])[1]>0.05, paired=False)[1]
		
		all_n_list.append(int(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == "combined", "N"].item()))
		all_mean_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == "combined", "Mean"].item())
		all_sd_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == "combined", "SD"].item())
		all_se_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == "combined", "SE"].item())
		all_ci_low_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == "combined", "95% Conf."].item())
		all_ci_high_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == "combined", "Interval"].item())
		group1_n_list.append(int(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == group1_label, "N"].item()))
		group1_mean_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == group1_label, "Mean"].item())
		group1_sd_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == group1_label, "SD"].item())
		group1_se_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == group1_label, "SE"].item())
		group1_ci_low_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == group1_label, "95% Conf."].item())
		group1_ci_high_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == group1_label, "Interval"].item())
		group2_n_list.append(int(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == group2_label, "N"].item()))
		group2_mean_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == group2_label, "Mean"].item())
		group2_sd_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == group2_label, "SD"].item())
		group2_se_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == group2_label, "SE"].item())
		group2_ci_low_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == group2_label, "95% Conf."].item())
		group2_ci_high_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == group2_label, "Interval"].item())
		equal_var_list.append(stats.levene(group1_df[var], group2_df[var])[1]>0.05)
		deg_free_list.append(ttest_stats_df2.at[1,ttest_stats_df2.columns[1]])
		t_stat_list.append(ttest_stats_df2.at[2,ttest_stats_df2.columns[1]])
		effectsize_list.append(ttest_stats_df2.at[effect_size_dfrow,ttest_stats_df2.columns[1]])
		pvalues_list.append(ttest_stats_df2.at[3,ttest_stats_df2.columns[1]])

	output_df = pd.DataFrame()
	output_df["Variable"] = variables_list
	output_df["All_N"] = all_n_list
	output_df["All_Mean"] = all_mean_list
	output_df["All_SD"] = all_sd_list
	output_df["All_Std Err"] = all_se_list
	output_df["All_95% CI_Low"] = all_ci_low_list
	output_df["All_95% CI_High"] = all_ci_high_list
	output_df[group1_label+"_N"] = group1_n_list
	output_df[group1_label+"_Mean"] = group1_mean_list
	output_df[group1_label+"_SD"] = group1_sd_list
	output_df[group1_label+"_Std Err"] = group1_se_list
	output_df[group1_label+"_95% CI_Low"] = group1_ci_low_list
	output_df[group1_label+"_95% CI_High"] = group1_ci_high_list
	output_df[group2_label+"_N"] = group2_n_list
	output_df[group2_label+"_Mean"] = group2_mean_list
	output_df[group2_label+"_SD"] = group2_sd_list
	output_df[group2_label+"_Std Err"] = group2_se_list
	output_df[group2_label+"_95% CI_Low"] = group2_ci_low_list
	output_df[group2_label+"_95% CI_High"] = group2_ci_high_list
	output_df["Equal Variances Assumed"] = equal_var_list
	output_df["Degrees of Freedom"] = deg_free_list
	output_df["t"] = t_stat_list
	output_df[effect_size_choice] = effectsize_list
	output_df["pvalues"] = pvalues_list
	
	output_df["Equal Variances Assumed"] = output_df["Equal Variances Assumed"].replace(True,'Yes')
	output_df["Equal Variances Assumed"] = output_df["Equal Variances Assumed"].replace(False,'No')

	return output_df
	
def raw_pairttest_generate_output_df(mod_raw_data_df):
	time1_n_list = []
	time1_mean_list = []
	time1_sd_list = []
	time1_se_list = []
	time1_ci_low_list = []
	time1_ci_high_list = []
	time2_n_list = []
	time2_mean_list = []
	time2_sd_list = []
	time2_se_list = []
	time2_ci_low_list = []
	time2_ci_high_list = []
	equal_var_list = []
	deg_free_list = []
	t_stat_list = []
	effectsize_list = []
	pvalues_list = []

	if effect_size_choice == "Cohen's d":
		effect_size_dfrow = 6
	elif effect_size_choice == "Hedge's g":
		effect_size_dfrow = 7
	elif effect_size_choice == "Glass's delta":
		effect_size_dfrow = 8

	for var_time1, var_time2 in raw_pairttest_inputVars_list:
		df_time1 = mod_raw_data_df[var_time1].dropna()
		df_time2 = mod_raw_data_df[var_time2].dropna()

		ttest_stats_df1 = rp.ttest(df_time1, df_time2, group1_name=var_time1, group2_name=var_time2, 
			 equal_variances=stats.levene(df_time1, df_time2)[1]>0.05, paired=True)[0]
		ttest_stats_df2 = rp.ttest(df_time1, df_time2, group1_name=var_time1, group2_name=var_time2, 
			 equal_variances=stats.levene(df_time1, df_time2)[1]>0.05, paired=True)[1]

		time1_n_list.append(int(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == var_time1, "N"].item()))
		time1_mean_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == var_time1, "Mean"].item())
		time1_sd_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == var_time1, "SD"].item())
		time1_se_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == var_time1, "SE"].item())
		time1_ci_low_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == var_time1, "95% Conf."].item())
		time1_ci_high_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == var_time1, "Interval"].item())
		time2_n_list.append(int(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == var_time2, "N"].item()))
		time2_mean_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == var_time2, "Mean"].item())
		time2_sd_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == var_time2, "SD"].item())
		time2_se_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == var_time2, "SE"].item())
		time2_ci_low_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == var_time2, "95% Conf."].item())
		time2_ci_high_list.append(ttest_stats_df1.loc[ttest_stats_df1["Variable"] == var_time2, "Interval"].item())
		equal_var_list.append(stats.levene(df_time1, df_time2)[1]>0.05)
		deg_free_list.append(ttest_stats_df2.at[1,ttest_stats_df2.columns[1]])
		t_stat_list.append(ttest_stats_df2.at[2,ttest_stats_df2.columns[1]])
		effectsize_list.append(ttest_stats_df2.at[effect_size_dfrow,ttest_stats_df2.columns[1]])
		pvalues_list.append(ttest_stats_df2.at[3,ttest_stats_df2.columns[1]])

	output_df = pd.DataFrame()
	output_df["Variable"] = ["{var1} - {var2}".format(var1=x, var2=y) for x,y in raw_pairttest_inputVars_list]
	output_df["Time1_N"] = time1_n_list
	output_df["Time1_Mean"] = time1_mean_list
	output_df["Time1_SD"] = time1_sd_list
	output_df["Time1_Std Err"] = time1_se_list
	output_df["Time1_95% CI_Low"] = time1_ci_low_list
	output_df["Time1_95% CI_High"] = time1_ci_high_list
	output_df["Time2_N"] = time2_n_list
	output_df["Time2_Mean"] = time2_mean_list
	output_df["Time2_SD"] = time2_sd_list
	output_df["Time2_Std Err"] = time2_se_list
	output_df["Time2_95% CI_Low"] = time2_ci_low_list
	output_df["Time2_95% CI_High"] = time2_ci_high_list
	output_df["Equal Variances Assumed"] = equal_var_list
	output_df["Degrees of Freedom"] = deg_free_list
	output_df["t"] = t_stat_list
	output_df[effect_size_choice] = effectsize_list
	output_df["pvalues"] = pvalues_list

	output_df["Equal Variances Assumed"] = output_df["Equal Variances Assumed"].replace(True,'Yes')
	output_df["Equal Variances Assumed"] = output_df["Equal Variances Assumed"].replace(False,'No')

	return output_df
def summ_corr_generate_output_df(mod_raw_data_df):
	#all checks are done so no need for modifications; function used only for infrastructure consitency
	output_df = mod_raw_data_df.copy()

	return output_df

def summ_indttest_generate_output_df(mod_raw_data_df):
	output_dict = {summ_indttest_var: [], summ_indttest_meanOne: [], summ_indttest_sdOne: [], summ_indttest_meanTwo: [],
					summ_indttest_sdTwo: [], "Degrees of Freedom": [], "t": [], effect_size_choice: [], "pvalues": []}

	for row in range(0, len(mod_raw_data_df)):
		current_df = mod_raw_data_df.loc[row]
		n1 = current_df[summ_indttest_nOne]
		n2 = current_df[summ_indttest_nTwo]

		output_dict[summ_indttest_var].append(current_df[summ_indttest_var])
		output_dict[summ_indttest_meanOne].append(current_df[summ_indttest_meanOne])
		output_dict[summ_indttest_sdOne].append(current_df[summ_indttest_sdOne])
		output_dict[summ_indttest_meanTwo].append(current_df[summ_indttest_meanTwo])
		output_dict[summ_indttest_sdTwo].append(current_df[summ_indttest_sdTwo])
		if current_df[summ_indttest_equal_var]: #equal vars assumed
			deg_free = n1 + n2 - 2
		else:
			s1_pooled = current_df[summ_indttest_sdOne]**2
			s2_pooled = current_df[summ_indttest_sdTwo]**2
			deg_free = (s1_pooled/n1 + s2_pooled/n2)**2 / ( (1/(n1-1)*(s1_pooled/n1)**2) + (1/(n2-1)*(s2_pooled/n2)**2) )
		output_dict["Degrees of Freedom"].append(deg_free)
		#stats.ttest_ind_from_stats itself decides whether to to perform student's t-test or welch's based on equal_var
		t, p = stats.ttest_ind_from_stats(mean1 = current_df[summ_indttest_meanOne], std1 = current_df[summ_indttest_sdOne], 
										nobs1 = n1, mean2 = current_df[summ_indttest_meanTwo], std2 = current_df[summ_indttest_sdTwo], 
										nobs2 = n2, equal_var=current_df[summ_indttest_equal_var])
		output_dict["t"].append(t)
		output_dict["pvalues"].append(p)
		
		if effect_size_choice == "Cohen's d":
			effect_size = t*np.sqrt(1/n1 + 1/n2)
		elif effect_size_choice == "Hedge's g":
			#calculated using cohens_d * (1 - (3/4(n1+n2-9))) where cohens d is t*sqrt(1/n1 + 1/n2)
			effect_size = (t*np.sqrt(1/n1 + 1/n2))*np.sqrt(1/n1 + 1/n2)
		elif effect_size_choice == "Glass's delta":
			#glass delta = (Mean2 - Mean1) / SD1 where SD1 is always that of control
			effect_size = (current_df[summ_indttest_meanTwo] - current_df[summ_indttest_meanOne]) / current_df[summ_indttest_sdOne]
		output_dict[effect_size_choice].append(effect_size)

	output_df = pd.DataFrame(output_dict)
	
	return output_df

def spss_corr_generate_output_df(mod_raw_data_df):
	correlation_label = mod_raw_data_df.iloc[0,1]
	mod_raw_data_df.drop(columns=mod_raw_data_df.columns[1], inplace=True)

	variables_list = list(mod_raw_data_df.columns)[1:]

	row_var_list = []
	col_var_list = []
	corr_coeff_list = []
	pvalues_list = []
	
	column = 2
	for row in range(0, len(variables_list)*2,2):
		for col in range(column, len(variables_list)+1):
			row_var_list.append(mod_raw_data_df.iloc[row,0])
			col_var_list.append(mod_raw_data_df.columns[col])
			corr_coeff_list.append(mod_raw_data_df.iloc[row,col])
			pvalues_list.append(mod_raw_data_df.iloc[row+1,col])
		column += 1

	output_df = pd.DataFrame()
	output_df["var1"] = row_var_list
	output_df["var2"] = col_var_list
	output_df[correlation_label] = corr_coeff_list
	output_df["pvalues"] = pvalues_list
	
	return output_df

def spss_mr_generate_output_df(mod_raw_data_df):
	variables_list = ["(Constant)"] + list(mod_raw_data_df[mod_raw_data_df.columns[1]])[1:-1]
	
	b_list = ["{:.2f}".format(x) for x in list(mod_raw_data_df["Unstandardized Coefficients"])[:-1]]
	try:
		ci_low_list = ["{:.2f}".format(x) for x in list(mod_raw_data_df["95.0% Confidence Interval for B"])[:-1]]
	except KeyError:
		ci_low_list = [""] * (len(variables_list))
	try:
		ci_high_list = ["{:.2f}".format(x) for x in list(mod_raw_data_df["Upper Bound"])[:-1]]
	except KeyError:
		ci_high_list = [""] * (len(variables_list))
	beta_list = ["{:.2f}".format(x) for x in list(mod_raw_data_df["Standardized Coefficients"])[:-1]]
	t_list = ["{:.2f}".format(x) for x in list(mod_raw_data_df["t"])[:-1]]
	pvalues_list = [pvalue_formatting(x) for x in list(mod_raw_data_df["Sig."])[:-1]] #can afford to format here as there will be no multitest corrections applied

	output_df = pd.DataFrame()
	output_df["Variable"] = variables_list
	output_df["B"] = b_list
	output_df["95% CI"] = ["["+low+","+high+"]" for low,high in zip(ci_low_list, ci_high_list)]
	output_df["beta"] = ['' if x=='nan' else x for x in beta_list] #the beta for the constant is missing from the spss output
	output_df["t"] = t_list
	output_df["p"] = pvalues_list
	
	return output_df

def spss_indttest_generate_output_df(mod_raw_data_df):
	variables_list = []
	levene_sig_list = []
	t_stat_list = []
	pvalues_list = []
	deg_free_list = []
	#ci_low_list = []
	#ci_high_list = []
	effectsize_list = []

	for row in range(1, len(mod_raw_data_df), 2):
		variables_list.append(mod_raw_data_df["Independent Samples Test"][row])
		levene_p = mod_raw_data_df["Sig."][row]
		if levene_p > 0.05:
			row_to_read = row
		else:
			row_to_read = row+1
		t = mod_raw_data_df["t"][row_to_read]
		t_stat_list.append(t)
		pvalues_list.append(mod_raw_data_df["Sig. (2-tailed)"][row_to_read])
		deg_free_list.append(mod_raw_data_df["df"][row_to_read])
		#ci_low_list.append(mod_raw_data_df[mod_raw_data_df.columns[9]][row_to_read])
		#ci_high_list.append(mod_raw_data_df[mod_raw_data_df.columns[10]][row_to_read])

		if effect_size_choice == "Cohen's d":
			effect_size = t*np.sqrt(1/spss_indttest_nOne + 1/spss_indttest_nTwo)
		elif effect_size_choice == "Hedge's g":
			#calculated using cohens_d * (1 - (3/4(n1+n2-9))) where cohens d is t*sqrt(1/n1 + 1/n2)
			effect_size = (t*np.sqrt(1/spss_indttest_nOne + 1/spss_indttest_nTwo))*np.sqrt(1/spss_indttest_nOne + 1/spss_indttest_nTwo)
		effectsize_list.append(effect_size)

	output_df = pd.DataFrame()
	output_df["Variable"] = variables_list
	output_df["All_Mean"] = [""] * len(output_df)
	output_df["All_SD"] = [""] * len(output_df)
	output_df[spss_indttest_groupOneLabel+"_Mean"] = [""] * len(output_df)
	output_df[spss_indttest_groupOneLabel+"_SD"] = [""] * len(output_df)
	output_df[spss_indttest_groupTwoLabel+"_Mean"] = [""] * len(output_df)
	output_df[spss_indttest_groupTwoLabel+"_SD"] = [""] * len(output_df)
	output_df["Degrees of Freedom"] = ["{:.2f}".format(x) for x in deg_free_list]
	output_df["t"] = ["{:.2f}".format(x) for x in t_stat_list]
	output_df[effect_size_choice] = ["{:.2f}".format(x) for x in effectsize_list]
	output_df["pvalues"] = pvalues_list

	return output_df
	
def spss_pairttest_generate_output_df(mod_raw_data_df):
	t_stat_list = list(mod_raw_data_df["t"])
	effectsize_list = []

	if effect_size_choice == "Cohen's d":
		for t in t_stat_list:
			effect_size = t*np.sqrt(1/spss_pairttest_nOne + 1/spss_pairttest_nTwo)
			effectsize_list.append(effect_size)
	elif effect_size_choice == "Hedge's g":
		#calculated using cohens_d * (1 - (3/4(n1+n2-9))) where cohens d is t*sqrt(1/n1 + 1/n2)
		for t in t_stat_list:
			effect_size = (t*np.sqrt(1/spss_pairttest_nOne + 1/spss_pairttest_nTwo))*np.sqrt(1/spss_pairttest_nOne + 1/spss_pairttest_nTwo)
			effectsize_list.append(effect_size)
			
	output_df = pd.DataFrame()
	output_df["Variables"] = list(mod_raw_data_df["Variables"])
	output_df["Time 1_Mean"] = [""] * len(output_df)
	output_df["Time 1_SD"] = [""] * len(output_df)
	output_df["Time 2_Mean"] = [""] * len(output_df)
	output_df["Time 2_SD"] = [""] * len(output_df)
	output_df["df"] = ["{:.2f}".format(x) for x in list(mod_raw_data_df["df"])]
	output_df["t"] = ["{:.2f}".format(x) for x in t_stat_list]
	output_df[effect_size_choice] = ["{:.2f}".format(x) for x in effectsize_list]
	output_df["pvalues"] = list(mod_raw_data_df["p"])

	return output_df

def pvalues_generate_output_df(mod_raw_data_df):
	output_df = mod_raw_data_df.copy()

	return output_df

#-----------------1.3. Individual functions for saving data
#1.3.1. Helper functions for saving data
alignment_top = Alignment(horizontal="center", vertical="top")
alignment_center = Alignment(horizontal="center", vertical="center")
font_title = Font(size=20, bold=True)
font_header = Font(italic=True)
font_bold = Font(bold=True)
border_APA = Side(border_style="medium", color="000000")

def pvalue_formatting(x):
	if x<0.001:
		return "<.001"
	else:
		return "{:.3f}".format(x).replace("0", "", 1)

def correlations_format_val(x, p=None):
	try:
		if isinstance(x, str): #covers the case when the correlations come from SPSS input and are flagged significant
			x = "{:.2f}".format(float("".join([char for char in x if char!="*"]))).replace("0","",1)
		elif abs(x) == 1.0:
			x = "{:.2f}".format(val)
		else:
			x = "{:.2f}".format(x).replace("0","",1)
	except:
		raise Exception("Something went wrong with formatting the correlation coefficients. Please try again.")
	if p != None:
		if p < 0.001:
			x = x + "**"
		elif p < 0.05:
			x = x + "*"
	return x

def add_table_notes(ws, table_notes, adjusted_pvalues_threshold=None):
	if correction_type != "":
		table_notes.append("Multiple tests correction applied to p values: {c}".format(c=list(correction_type_Dict.keys())[list(correction_type_Dict.values()).index(correction_type)]))
		if output_pvalues_type == "adjusted":
			table_notes.append("Adjusted p values used; significant at {alpha}".format(alpha=alpha_threshold))
		elif output_pvalues_type == "original":
			table_notes.append("Original p values used; significant at {alpha}".format(alpha=adjusted_pvalues_threshold))
		
	for note in table_notes:
		ws.append([note])

def save_file(output_name, wb):
	time_now = datetime.now().strftime("%H%M%S-%Y%m%d")
	filename = output_dir + "/" + output_name + "_" + time_now + ".xlsx"
	wb.save(filename=filename)

#1.3.2. Individual functions for saving data
def raw_corr_apa_table(mod_raw_data_df, output_df):
	wb = Workbook()
	ws = wb.active
	
	variables_list = list(mod_raw_data_df.columns)
	
	ws.append([""] + variables_list[:-1] + ["Mean", "SD"])
	ws["A2"] = variables_list[0]
	for ind,var in enumerate(variables_list[1:]):
		ws.cell(row=(ind*2)+3, column=1).value = var
	
	start_row = 3
	for col in range(2, len(variables_list)+3):
		col_var = ws.cell(row=1, column=col).value
		for row in range(start_row, len(variables_list)*2, 2):
			row_var = ws.cell(row=row, column=1).value
			#query method is slightly slower for smaller datasets
			r = output_df[((output_df["Variable1"]==row_var) & (output_df["Variable2"]==col_var)) | ((output_df["Variable1"]==col_var) & (output_df["Variable2"]==row_var))]["Correlation Coefficient"].item()
			p = output_df[((output_df["Variable1"]==row_var) & (output_df["Variable2"]==col_var)) | ((output_df["Variable1"]==col_var) & (output_df["Variable2"]==row_var))]["adjusted_pvalues"].item()
			ci_low = output_df[((output_df["Variable1"]==row_var) & (output_df["Variable2"]==col_var)) | ((output_df["Variable1"]==col_var) & (output_df["Variable2"]==row_var))]["CI_low"].item()
			ci_high = output_df[((output_df["Variable1"]==row_var) & (output_df["Variable2"]==col_var)) | ((output_df["Variable1"]==col_var) & (output_df["Variable2"]==row_var))]["CI_high"].item()
			#r = output_df.query("(Variable1 == @row_var and Variable2==@col_var) or (Variable1 == @col_var and Variable2==@row_var)")["Correlation Coefficient"].item()
			#p = output_df.query("(Variable1 == @row_var and Variable2==@col_var) or (Variable1 == @col_var and Variable2==@row_var)")["adjusted_pvalues"].item()
			#ci_low = output_df.query("(Variable1 == @row_var and Variable2==@col_var) or (Variable1 == @col_var and Variable2==@row_var)")["CI_low"].item()
			#ci_high = output_df.query("(Variable1 == @row_var and Variable2==@col_var) or (Variable1 == @col_var and Variable2==@row_var)")["CI_high"].item()
			r, ci_low, ci_high = correlations_format_val(r, p), correlations_format_val(ci_low), correlations_format_val(ci_high)
			ws.cell(row=row, column=col).value = r
			ws.cell(row=row+1, column=col).value = "[" + ci_low + ", " + ci_high + "]"
		start_row += 2
	
	ws.cell(row=2, column=len(variables_list)+1).value = "{:.2f}".format(mod_raw_data_df[variables_list[0]].mean(skipna=True))
	ws.cell(row=2, column=len(variables_list)+2).value = "{:.2f}".format(mod_raw_data_df[variables_list[0]].std(skipna=True))
	for row in range(3, len(variables_list)*2, 2):
		ws.cell(row=row, column=len(variables_list)+1).value = "{:.2f}".format(mod_raw_data_df[variables_list[int((row-1)/2)]].mean(skipna=True))
		ws.cell(row=row, column=len(variables_list)+2).value = "{:.2f}".format(mod_raw_data_df[variables_list[int((row-1)/2)]].std(skipna=True))    
	
	for row in range(3, len(variables_list)*2, 2):
		ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=1)
		ws.merge_cells(start_row=row, start_column=len(variables_list)+1, end_row=row+1, end_column=len(variables_list)+1)
		ws.merge_cells(start_row=row, start_column=len(variables_list)+2, end_row=row+1, end_column=len(variables_list)+2)
		
	for row in ws[1:len(variables_list)*2]:
		for cell in row:
			if cell.row > 2 and (cell.column == 1 or cell.column == len(variables_list)+1 or cell.column == len(variables_list)+2):
				cell.alignment = alignment_top
			else:
				cell.alignment = alignment_center
				
	for cell in ws[1]:
		cell.font = font_header
	
	ci_font = Font(size=9)
	for row in range(4, len(variables_list)*2+1, 2):
		for cell in ws[row]:
			cell.font = ci_font
	
	for cell in ws[1]:
		cell.border = Border(top=border_APA, bottom=border_APA)
	for cell in ws[len(variables_list)*2]:
		cell.border = Border(bottom = border_APA)
	
	table_notes = ["Note. M and SD represent mean and standard deviation, respectively. Values in square brackets indicate 95% confidence interval for the correlation coefficient."]
	table_notes.append("Correlation coefficient used: {}".format(raw_corr_type))
	table_notes.append("**p < 0.01")
	table_notes.append("*p < {}".format(alpha_threshold))
	add_table_notes(ws, table_notes) 
	
	save_file("raw_data_correlations", wb)

def raw_mr_apa_table(mod_raw_data_df, output_df):
	mod_output_df = output_df[["Variable","B","95% CI B","beta","t","pvalues"]]

	mod_output_df["pvalues"] = mod_output_df["pvalues"].map(pvalue_formatting)

	mod_output_df.rename(columns = {"95% CI B": "95% CI", "pvalues": "p"}, inplace=True)
	mod_output_df.loc[0, "beta"] = "" #removes the beta value for constant as it is always 0

	wb = Workbook()
	ws = wb.active

	ws.append(list(mod_output_df.columns))
	for row in dataframe_to_rows(mod_output_df, index=False, header=False):
		ws.append(row)

	for cell in ws[1]:
		cell.font = font_header

	for cell in ws[1] + ws[len(mod_output_df)+1]:
		cell.border = Border(top=border_APA, bottom=border_APA)
	for cell in ws[len(mod_output_df)+1]:
		cell.border = Border(bottom=border_APA)

	for row in range(1, len(mod_output_df)+2):
		for cell in ws[row]:
			cell.alignment = alignment_center

	table_notes = ["R squared adjusted = {R}".format(R=output_df["R2adj"][0]), "Dependent Variable: {DV}".format(DV=output_df["Dependent Variable"][0])]
	add_table_notes(ws, table_notes)

	save_file("raw_data_MR", wb)

def raw_indttest_apa_table(mod_raw_data_df, output_df):
	group1_label = list(output_df.columns)[7][:-2]
	group2_label = list(output_df.columns)[13][:-2]
	sign_bool_label = list(output_df.columns)[-1]
	
	apa_output_df = output_df[["Variable","All_Mean","All_SD", group1_label+"_Mean", group1_label+"_SD", group2_label+"_Mean", 
							group2_label+"_SD", "Degrees of Freedom", "t", effect_size_choice, "pvalues", 
							"adjusted_pvalues"]]

	if output_pvalues_type == "adjusted":
		adjusted_pvalues_threshold = None
		apa_output_df["pvalues"] = apa_output_df["adjusted_pvalues"]
	else:
		adjusted_pvalues_threshold = sign_bool_label[sign_bool_label.find("=")+2:]
	apa_output_df.drop(columns = ["adjusted_pvalues"], inplace=True)

	apa_output_df[list(apa_output_df.columns)[1:-1]] = apa_output_df[list(apa_output_df.columns)[1:-1]].applymap(lambda x: "{:.2f}".format(x))
	apa_output_df["pvalues"] = apa_output_df["pvalues"].map(pvalue_formatting)

	wb = Workbook()
	ws = wb.active
	
	ws.cell(row=1, column=1).value = "Variable"
	ws.merge_cells('A1:A2')
	ws.cell(row=1, column=1).font = font_header
	
	ws.cell(row=1, column=2).value = "All, n={}".format(output_df.iloc[0, 1])
	ws.merge_cells('B1:C1')
	
	ws.cell(row=1, column=4).value = "{g}, n={n}".format(g=group1_label, n=output_df.iloc[0,7])
	ws.merge_cells('D1:E1')
	
	ws.cell(row=1, column=6).value = "{g}, n={n}".format(g=group2_label, n=output_df.iloc[0,13])
	ws.merge_cells('F1:G1')
	
	ws.cell(row=1, column=8).value = "df"
	ws.merge_cells('H1:H2')
	ws.cell(row=1, column=8).font = font_header
	
	ws.cell(row=1, column=9).value = "t"
	ws.merge_cells('I1:I2')
	ws.cell(row=1, column=9).font = font_header
	
	ws.cell(row=1, column=10).value = effect_size_choice
	ws.merge_cells('J1:J2')
	ws.cell(row=1, column=10).font = font_header
	
	ws.cell(row=1, column=11).value = "p"
	ws.merge_cells('K1:K2')
	ws.cell(row=1, column=11).font = font_header

	for col in range(2,7,2):
		ws.cell(row=2, column=col).value = "M"
		ws.cell(row=2, column=col).font = font_header
		ws.cell(row=2, column=col+1).value = "SD"
		ws.cell(row=2, column=col+1).font = font_header

	for row in dataframe_to_rows(apa_output_df, index=False, header=False):
		ws.append(row)

	for row in range(1, len(apa_output_df)+3):
		for cell in ws[row]:
			cell.alignment = alignment_center

	for cell in ws[1]:
		cell.border = Border(top=border_APA)
	for cell in ws[2] + ws[len(apa_output_df)+2]:
		cell.border = Border(bottom=border_APA)

	add_table_notes(ws, [], adjusted_pvalues_threshold)

	save_file("raw_data_indttest", wb)
					   
def raw_indttest_descriptives_table(mod_raw_data_df, output_df):
	if correction_type == "":
		output_df.drop(columns = ["adjusted_pvalues", list(output_df.columns)[-1]], inplace=True)
	
	test_stats_df = output_df[list(output_df.columns)[0:1] + list(output_df.columns)[19:]]
	descriptive_stats_df = output_df.drop(columns = list(output_df.columns)[19:])
	
	wb = Workbook()
	ws = wb.active
	
	ws.cell(row=1, column=1).value = "Descriptives"
	ws.cell(row=1, column=1).font = font_title
	ws.merge_cells('A1:S1')
	
	for row in dataframe_to_rows(descriptive_stats_df, index=False, header=True):
		ws.append(row)
	
	for cell in ws[2]:
		cell.font = font_bold
		
	ws.cell(row=len(descriptive_stats_df)+5,column=1).value = "Test Statistics"
	ws.cell(row=len(descriptive_stats_df)+5,column=1).font = font_title
	ws.merge_cells(start_row=len(descriptive_stats_df)+5, start_column=1, end_row=len(descriptive_stats_df)+5, end_column=len(test_stats_df.columns))
	
	for row in dataframe_to_rows(test_stats_df, index=False, header=True):
		ws.append(row)

	for cell in ws[len(descriptive_stats_df)+6]:
		cell.font = font_bold

	for row in range(1, len(descriptive_stats_df) + len(test_stats_df) + 7):
		for cell in ws[row]:
			cell.alignment = alignment_center

	save_file("raw_data_indttest_descriptives", wb)

def raw_pairttest_apa_table(mod_raw_data_df, output_df):
	sign_bool_label = list(output_df.columns)[-1]
	
	apa_output_df = output_df[["Variable", "Time1_Mean", "Time1_SD", "Time2_Mean", "Time2_SD"] + list(output_df.columns)[14:-1]]

	if output_pvalues_type == "adjusted":
		adjusted_pvalues_threshold = None
		apa_output_df["pvalues"] = apa_output_df["adjusted_pvalues"]
	else:
		adjusted_pvalues_threshold = sign_bool_label[sign_bool_label.find("=")+2:]
	apa_output_df.drop(columns = ["adjusted_pvalues"], inplace=True)

	apa_output_df[list(apa_output_df.columns)[1:-1]] = apa_output_df[list(apa_output_df.columns)[1:-1]].applymap(lambda x: "{:.2f}".format(x))
	apa_output_df["pvalues"] = apa_output_df["pvalues"].map(pvalue_formatting)

	wb = Workbook()
	ws = wb.active
	
	ws.cell(row=1, column=1).value = "Variable"
	ws.merge_cells('A1:A2')
	ws.cell(row=1, column=1).font = font_header
	
	ws.cell(row=1, column=2).value = "Time 1"
	ws.merge_cells('B1:C1')
	
	ws.cell(row=1, column=4).value = "Time 2"
	ws.merge_cells('D1:E1')
	
	ws.cell(row=1, column=6).value = "df"
	ws.merge_cells('F1:F2')
	ws.cell(row=1, column=6).font = font_header
	
	ws.cell(row=1, column=7).value = "t"
	ws.merge_cells('G1:G2')
	ws.cell(row=1, column=7).font = font_header
	
	ws.cell(row=1, column=8).value = effect_size_choice
	ws.merge_cells('H1:H2')
	ws.cell(row=1, column=8).font = font_header
	
	ws.cell(row=1, column=9).value = "p"
	ws.merge_cells('I1:I2')
	ws.cell(row=1, column=9).font = font_header
	
	for col in range(2,5,2):
		ws.cell(row=2, column=col).value = "M"
		ws.cell(row=2, column=col).font = font_header
		ws.cell(row=2, column=col+1).value = "SD"
		ws.cell(row=2, column=col+1).font = font_header
	
	for row in dataframe_to_rows(apa_output_df, index=False, header=False):
		ws.append(row)

	for row in range(1, len(apa_output_df)+3):
		for cell in ws[row]:
			cell.alignment = alignment_center
	
	for cell in ws[2] + ws[len(apa_output_df)+2]:
		cell.border = Border(bottom=border_APA)

	for cell in ws[1]:
		cell.border = Border(top=border_APA)

	add_table_notes(ws, [], adjusted_pvalues_threshold)

	save_file("raw_data_pairttest", wb)
	
def raw_pairttest_descriptives_table(mod_raw_data_df, output_df):
	if correction_type == "":
		output_df.drop(columns = ["adjusted_pvalues"], inplace=True)

	descriptive_stats_df = output_df[list(output_df.columns)[0:13]]
	test_stats_df = output_df[list(output_df.columns)[0:1] + list(output_df.columns)[13:]]

	wb = Workbook()
	ws = wb.active

	ws.cell(row=1, column=1).value = "Descriptives"
	ws.cell(row=1, column=1).font = font_title
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(descriptive_stats_df.columns))

	for row in dataframe_to_rows(descriptive_stats_df, index=False, header=True):
		ws.append(row)

	ws.cell(row=len(descriptive_stats_df)+5,column=1).value = "Test Statistics"
	ws.cell(row=len(descriptive_stats_df)+5,column=1).font = font_title
	ws.merge_cells(start_row=len(descriptive_stats_df)+5, start_column=1,
				   end_row=len(descriptive_stats_df)+5, end_column=len(test_stats_df.columns))
	
	for row in dataframe_to_rows(test_stats_df, index=False, header=True):
		ws.append(row)

	for cell in ws[2] + ws[len(test_stats_df)+6]:
		cell.font = font_header
		
	for row in range(1, len(descriptive_stats_df) + len(test_stats_df) + 7):
		for cell in ws[row]:
			cell.alignment = alignment_center

	save_file("raw_data_pairttest_descriptives", wb)

def summ_corr_apa_table(mod_raw_data_df, output_df):
	variables_list = list(output_df[summ_corr_varOne].value_counts().keys()) + [list(output_df[summ_corr_varTwo].value_counts().keys())[0]]

	wb = Workbook()
	ws = wb.active

	ws.append([""] + variables_list[1:])
	for ind,var in enumerate(variables_list[:-1]):
		ws.cell(row=ind+2, column=1).value = var

	start_col = 2
	for row in range(2, len(variables_list)+1):
		row_var = ws.cell(row=row, column=1).value
		for col in range(start_col, len(variables_list)+1):
			col_var = ws.cell(row=1, column=col).value
			#here query method is not only slower as it is much smaller dataset but also cannot refer to two different vaiables (colname and val)
			r = output_df[((output_df[summ_corr_varOne]==row_var) & (output_df[summ_corr_varTwo]==col_var)) | ((output_df[summ_corr_varOne]==col_var) & (output_df[summ_corr_varTwo]==row_var))][summ_corr_coeff].item()
			p = output_df[((output_df[summ_corr_varOne]==row_var) & (output_df[summ_corr_varTwo]==col_var)) | ((output_df[summ_corr_varOne]==col_var) & (output_df[summ_corr_varTwo]==row_var))]["adjusted_pvalues"].item()
			r = correlations_format_val(r, p)
			ws.cell(row=row, column=col).value = r
		start_col += 1

	for row in range(1, len(variables_list)+1):
		for cell in ws[row]:
			cell.alignment = alignment_center

	for cell in ws[1]:
		cell.font = font_header

	for row in range(2, len(variables_list)+1):
		ws.cell(row=row, column=1).font = font_header

	for cell in ws[len(variables_list)]:
		cell.border = Border(bottom=border_APA)

	for cell in ws[1]:
		cell.border = Border(top=border_APA, bottom=border_APA)

	table_notes = ["**p < 0.01", "*p < {}".format(alpha_threshold)]
	add_table_notes(ws, table_notes)

	save_file("summ_correlations", wb)
	
def summ_indttest_apa_table(mod_raw_data_df, output_df):
	sign_bool_label = list(output_df.columns)[-1]

	if output_pvalues_type == "adjusted":
		adjusted_pvalues_threshold = None
		output_df["pvalues"] = output_df["adjusted_pvalues"]
	else:
		adjusted_pvalues_threshold = sign_bool_label[sign_bool_label.find("=")+2:]
	output_df.drop(columns = ["adjusted_pvalues", sign_bool_label], inplace=True)

	output_df[list(output_df.columns)[1:-1]] = output_df[list(output_df.columns)[1:-1]].applymap(lambda x: "{:.2f}".format(x))
	output_df["pvalues"] = output_df["pvalues"].map(pvalue_formatting)

	wb = Workbook()
	ws = wb.active

	ws.cell(row=1, column=1).value = "Variable"
	ws.merge_cells('A1:A2')
	ws.cell(row=1, column=1).font = font_header
	
	ws.cell(row=1, column=2).value = "Group1, n={n}".format(n=mod_raw_data_df[summ_indttest_nOne][0])
	ws.merge_cells('B1:C1')
	
	ws.cell(row=1, column=4).value = "Group2, n={n}".format(n=mod_raw_data_df[summ_indttest_nTwo][0])
	ws.merge_cells('D1:E1')
	
	ws.cell(row=1, column=6).value = "df"
	ws.merge_cells('F1:F2')
	ws.cell(row=1, column=6).font = font_header
	
	ws.cell(row=1, column=7).value = "t"
	ws.merge_cells('G1:G2')
	ws.cell(row=1, column=7).font = font_header
	
	ws.cell(row=1, column=8).value = effect_size_choice
	ws.merge_cells('H1:H2')
	ws.cell(row=1, column=8).font = font_header
	
	ws.cell(row=1, column=9).value = "p"
	ws.merge_cells('I1:I2')
	ws.cell(row=1, column=9).font = font_header
	
	for col in range(2,5,2):
		ws.cell(row=2, column=col).value = "M"
		ws.cell(row=2, column=col).font = font_header
		ws.cell(row=2, column=col+1).value = "SD"
		ws.cell(row=2, column=col+1).font = font_header
	
	for row in dataframe_to_rows(output_df, index=False, header=False):
		ws.append(row)

	for row in range(1, len(output_df)+3):
		for cell in ws[row]:
			cell.alignment = alignment_center

	for cell in ws[1]:
		cell.border = Border(top=border_APA)
	for cell in ws[2] + ws[len(output_df)+2]:
		cell.border = Border(bottom=border_APA)

	add_table_notes(ws, [], adjusted_pvalues_threshold)

	save_file("summ_indttest", wb)

def spss_corr_apa_table(mod_raw_data_df, output_df):
	print("Function executed: spss_corr_apa_table")
	#output_df.to_excel('asdasd3.xlsx')
	correlation_label = output_df.columns[2]
	variables_list = mod_raw_data_df.columns[1:]

	apa_table_df = pd.DataFrame(index = variables_list[:-1], columns = variables_list[1:])

	column = 0
	for row in range(0, len(apa_table_df)):
		row_var = apa_table_df.index[row]
		for col in range(column, len(apa_table_df.columns)):
			col_var = apa_table_df.columns[col]
			r = output_df[((output_df["var1"]==row_var) & (output_df["var2"]==col_var)) | ((output_df["var1"]==col_var) & (output_df["var2"]==row_var))][correlation_label].item()
			p = output_df[((output_df["var1"]==row_var) & (output_df["var2"]==col_var)) | ((output_df["var1"]==col_var) & (output_df["var2"]==row_var))]["adjusted_pvalues"].item()
			apa_table_df.iloc[row, col] = correlations_format_val(r, p)
		column += 1

	wb = Workbook()
	ws = wb.active

	for row in dataframe_to_rows(apa_table_df, index=True, header=True):
		ws.append(row)
	ws.delete_rows(2) #the second row in the spreadsheet is blank for some reason so we get rid of it

	for cell in ws['A'] + ws[1]:
		cell.font = font_header
		
	for col in range(1, len(apa_table_df.columns)+2):
		for row in range(1, len(apa_table_df)+2):
			ws.cell(row=row, column=col).alignment = alignment_center
	
	for cell in ws[1]:
		cell.border = Border(top=border_APA, bottom=border_APA)

	for cell in ws[len(apa_table_df)+1]:
		cell.border = Border(bottom=border_APA)

	table_notes = ["Correlation coefficient used: {}".format(correlation_label), "**p < 0.01", "*p < {}".format(alpha_threshold)]
	add_table_notes(ws, table_notes)

	save_file("spss_correlations", wb)

def spss_mr_apa_table(mod_raw_data_df, output_df):
	wb = Workbook()
	ws = wb.active

	for row in dataframe_to_rows(output_df, index=False, header=True):
		ws.append(row)

	for cell in ws[1]:
		cell.font = font_header
		
	for row in range(1, len(output_df)+2):
		for cell in ws[row]:
			cell.alignment = alignment_center
	
	for cell in ws[1]:
		cell.border = Border(bottom=border_APA, top=border_APA)

	for cell in ws[len(output_df)+1]:
		cell.border = Border(bottom=border_APA)

	DV_cell = mod_raw_data_df[mod_raw_data_df.columns[0]][len(mod_raw_data_df[mod_raw_data_df.columns[0]])-1]
	DV = DV_cell[DV_cell.find(":")+2:]
	table_notes = ["R squared adjusted = X.XX", "Dependent variable: {}".format(DV)]
	if output_df["95% CI"][0] == "[,]":
		table_notes.append("Confidence intervals were not found in the SPSS table. Please add them to your SPSS table and re-run the program or add them manually.")

	add_table_notes(ws, table_notes)

	save_file("spss_mr", wb)

def spss_indttest_apa_table(mod_raw_data_df, output_df):
	apa_output_df = output_df.copy()

	if output_pvalues_type == "adjusted":
		adjusted_pvalues_threshold = None
		apa_output_df["pvalues"] = apa_output_df["adjusted_pvalues"]
	else:
		adjusted_pvalues_threshold = apa_output_df.columns[-1][apa_output_df.columns[-1].find("=")+2:]

	apa_output_df.drop(columns=apa_output_df.columns[-2:], inplace=True)

	apa_output_df["pvalues"] = apa_output_df["pvalues"].map(pvalue_formatting)

	wb = Workbook()
	ws = wb.active
	
	ws.cell(row=1, column=1).value = "Variable"
	ws.merge_cells('A1:A2')
	ws.cell(row=1, column=1).font = font_header
	
	ws.cell(row=1, column=2).value = "All, n=?"
	ws.merge_cells('B1:C1')
	
	ws.cell(row=1, column=4).value = "{}, n=?".format(spss_indttest_groupOneLabel)
	ws.merge_cells('D1:E1')
	
	ws.cell(row=1, column=6).value = "{}, n=?".format(spss_indttest_groupTwoLabel)
	ws.merge_cells('F1:G1')
	
	ws.cell(row=1, column=8).value = "df"
	ws.merge_cells('H1:H2')
	ws.cell(row=1, column=8).font = font_header
	
	ws.cell(row=1, column=9).value = "t"
	ws.merge_cells('I1:I2')
	ws.cell(row=1, column=9).font = font_header
	
	ws.cell(row=1, column=10).value = effect_size_choice
	ws.merge_cells('J1:J2')
	ws.cell(row=1, column=10).font = font_header
	
	ws.cell(row=1, column=11).value = "p"
	ws.merge_cells('K1:K2')
	ws.cell(row=1, column=11).font = font_header

	for col in range(2,7,2):
		ws.cell(row=2, column=col).value = "M"
		ws.cell(row=2, column=col).font = font_header
		ws.cell(row=2, column=col+1).value = "SD"
		ws.cell(row=2, column=col+1).font = font_header

	for row in dataframe_to_rows(apa_output_df, index=False, header=False):
		ws.append(row)

	for row in range(1, len(apa_output_df)+3):
		for cell in ws[row]:
			cell.alignment = alignment_center

	for cell in ws[1]:
		cell.border = Border(top=border_APA)
	for cell in ws[2] + ws[len(apa_output_df)+2]:
		cell.border = Border(bottom=border_APA)

	table_notes = ["Means and Standard Deviations could not be read from the SPSS table. Please add them yourself or remove those columns if they are not needed."]
	add_table_notes(ws, table_notes, adjusted_pvalues_threshold)

	save_file("spss_data_indttest", wb)

def spss_pairttest_apa_table(mod_raw_data_df, output_df):
	apa_output_df = output_df.copy()

	if output_pvalues_type == "adjusted":
		adjusted_pvalues_threshold = None
		apa_output_df["pvalues"] = apa_output_df["adjusted_pvalues"]
	else:
		adjusted_pvalues_threshold = apa_output_df.columns[-1][apa_output_df.columns[-1].find("=")+2:]

	apa_output_df.drop(columns=apa_output_df.columns[-2:], inplace=True)

	apa_output_df["pvalues"] = apa_output_df["pvalues"].map(pvalue_formatting)

	wb = Workbook()
	ws = wb.active
	
	ws.cell(row=1, column=1).value = "Variable"
	ws.merge_cells('A1:A2')
	ws.cell(row=1, column=1).font = font_header
	
	ws.cell(row=1, column=2).value = "Time 1"
	ws.merge_cells('B1:C1')
	
	ws.cell(row=1, column=4).value = "Time 2"
	ws.merge_cells('D1:E1')
	
	ws.cell(row=1, column=6).value = "df"
	ws.merge_cells('F1:F2')
	ws.cell(row=1, column=6).font = font_header
	
	ws.cell(row=1, column=7).value = "t"
	ws.merge_cells('G1:G2')
	ws.cell(row=1, column=7).font = font_header
	
	ws.cell(row=1, column=8).value = effect_size_choice
	ws.merge_cells('H1:H2')
	ws.cell(row=1, column=8).font = font_header
	
	ws.cell(row=1, column=9).value = "p"
	ws.merge_cells('I1:I2')
	ws.cell(row=1, column=9).font = font_header
	
	for col in range(2,5,2):
		ws.cell(row=2, column=col).value = "M"
		ws.cell(row=2, column=col).font = font_header
		ws.cell(row=2, column=col+1).value = "SD"
		ws.cell(row=2, column=col+1).font = font_header
	
	for row in dataframe_to_rows(apa_output_df, index=False, header=False):
		ws.append(row)

	for row in range(1, len(apa_output_df)+3):
		for cell in ws[row]:
			cell.alignment = alignment_center
	
	for cell in ws[2] + ws[len(apa_output_df)+2]:
		cell.border = Border(bottom=border_APA)

	table_notes = ["Means and Standard Deviations could not be read from the SPSS table. Please add them yourself or remove those columns if they are not needed."]
	add_table_notes(ws, table_notes, adjusted_pvalues_threshold)

	save_file("spss_data_pairttest", wb)

def pvalues_table(mod_raw_data_df, output_df):
	wb = Workbook()
	ws = wb.active

	for row in dataframe_to_rows(output_df, index=False, header=True):
		ws.append(row)

	for row in range(1, len(output_df)+2):
		for cell in ws[row]:
			cell.alignment = alignment_center

	for cell in ws[1]:
		cell.font = font_bold

	for cell in ws[1] + ws[len(output_df)+1]:
		cell.border = Border(bottom=border_APA)

	add_table_notes(ws, [])

	save_file("pvalues", wb)

#-----------------1.4. Decision functions for generating and saving output
def get_raw_data_df():
	try:
		if input_fileext == ".xlsx":
			raw_data_df = pd.read_excel(input_filename+".xlsx", sheet_name=0)
		elif input_fileext == ".csv":
			raw_data_df = pd.read_csv(input_filename+".csv")

		return raw_data_df
	except FileNotFoundError:
		raise Exception("There is no file called \"{f}\" in the directory".format(f=input_filename+input_fileext))

def modify_raw_data_df(raw_data_df):
	if input_type == "raw":
		if raw_test == "pairttest":
			raw_input_pairttest_colNames_check(raw_data_df)
		numeric_cols = get_numeric_cols(raw_data_df)
		if non_numeric_input_raise_errors == True:
			error_on_non_numeric_input(raw_data_df, numeric_cols)
		mod_raw_data_df = raw_input_generate_mod_raw_data_df(raw_data_df, numeric_cols)
	elif input_type == "summ_corr" or input_type == "summ_indttest":
		provided_cols = ["pvalues", summ_corr_varOne, summ_corr_varTwo, summ_corr_coeff] if input_type == "summ_corr" else [summ_indttest_var, summ_indttest_meanOne, summ_indttest_sdOne, summ_indttest_nOne, summ_indttest_meanTwo, summ_indttest_sdTwo, summ_indttest_nTwo, summ_indttest_equal_var]
		summ_input_colNames_check(raw_data_df, provided_cols)
		numeric_cols = get_numeric_cols(raw_data_df)
		error_on_non_numeric_input(raw_data_df, numeric_cols)
		mod_raw_data_df = summ_input_generate_mod_raw_data_df(raw_data_df)
	elif input_type == "spss":
		if spss_test == "corr":
			mod_raw_data_df = spss_corr_generate_mod_raw_data_df(raw_data_df) #might remove the function depending on whether I add logical checks
		elif spss_test == "mr":
			mod_raw_data_df = spss_mr_generate_mod_raw_data_df(raw_data_df) #ditto
		elif spss_test == "indttest":
			mod_raw_data_df = spss_indttest_generate_mod_raw_data_df(raw_data_df)
		elif spss_test == "pairttest":
			mod_raw_data_df = spss_pairttest_generate_mod_raw_data_df(raw_data_df)
	elif input_type == "pvalues":
		if "pvalues" not in raw_data_df.columns:
			raise Exception("Column \'pvalues\' is not found in the provided file. Please make sure it is typed correctly.")
		error_on_non_numeric_input(raw_data_df, ["pvalues"])
		mod_raw_data_df = raw_data_df.copy()

	return mod_raw_data_df

def generate_output_df(mod_raw_data_df):
	if input_type == "raw":
		if raw_test == "corr":
			output_df = raw_corr_generate_output_df(mod_raw_data_df)
		elif raw_test == "mr":
			output_df = raw_mr_generate_output_df(mod_raw_data_df)
		elif raw_test == "indttest":
			output_df = raw_indttest_generate_output_df(mod_raw_data_df)
		elif raw_test == "pairttest":
			output_df = raw_pairttest_generate_output_df(mod_raw_data_df)
	elif input_type == "summ_corr":
		output_df = summ_corr_generate_output_df(mod_raw_data_df)
	elif input_type == "summ_indttest":
		output_df = summ_indttest_generate_output_df(mod_raw_data_df)
	elif input_type == "spss":
		if spss_test == "corr":
			output_df = spss_corr_generate_output_df(mod_raw_data_df)
		elif spss_test == "mr":
			output_df = spss_mr_generate_output_df(mod_raw_data_df)
		elif spss_test == "indttest":
			output_df = spss_indttest_generate_output_df(mod_raw_data_df)
		elif spss_test == "pairttest":
			output_df = spss_pairttest_generate_output_df(mod_raw_data_df)
	elif input_type == "pvalues":
		output_df = pvalues_generate_output_df(mod_raw_data_df)

	return output_df

def multitest_correction(output_df):
	pvalues_list = output_df["pvalues"]
	sign_col_label = "Adjusted p value significant at alpha = {alpha}".format(alpha=alpha_threshold)

	if correction_type != "":    
		if correction_type == "sidak":
			sign_col_label = "Original p value significant at corrected alpha = {alpha}".format(alpha=round(multitest.multipletests(pvalues_list, alpha=alpha_threshold, method=correction_type, is_sorted=False, returnsorted=False)[2],5))
		elif correction_type == "bonferroni":
			sign_col_label = "Original p value significant at corrected alpha = {alpha}".format(alpha=round(multitest.multipletests(pvalues_list, alpha=alpha_threshold, method=correction_type, is_sorted=False, returnsorted=False)[3],5))

		adjusted_pvalues = multitest.multipletests(pvalues_list, alpha=alpha_threshold, method=correction_type, is_sorted=False, returnsorted=False)[1]
		sign_bools = multitest.multipletests(pvalues_list, alpha=alpha_threshold, method=correction_type, is_sorted=False, returnsorted=False)[0]
	else:
		adjusted_pvalues = pvalues_list
		sign_bools = [bool(x < alpha_threshold) for x in pvalues_list]
		
	output_df["adjusted_pvalues"] = adjusted_pvalues
	output_df[sign_col_label] = sign_bools

	output_df[sign_col_label] = output_df[sign_col_label].replace(True,"Significant")
	output_df[sign_col_label] = output_df[sign_col_label].replace(False,"Non-significant")

	return output_df

def save_output(mod_raw_data_df, output_df):
	if input_type == "raw":
		if raw_test == "corr":
			raw_corr_apa_table(mod_raw_data_df, output_df)
		elif raw_test == "mr":
			raw_mr_apa_table(mod_raw_data_df, output_df)
		elif raw_test == "indttest":
			raw_indttest_apa_table(mod_raw_data_df, output_df)
			if raw_ttest_output_descriptives == True:
				raw_indttest_descriptives_table(mod_raw_data_df, output_df)
		elif raw_test == "pairttest":
			raw_pairttest_apa_table(mod_raw_data_df, output_df)
			if raw_ttest_output_descriptives == True:
				raw_pairttest_descriptives_table(mod_raw_data_df, output_df)
	elif input_type == "summ_corr":
		summ_corr_apa_table(mod_raw_data_df, output_df)
	elif input_type == "summ_indttest":
		summ_indttest_apa_table(mod_raw_data_df, output_df)
	elif input_type == "spss":
		if spss_test == "corr":
			spss_corr_apa_table(mod_raw_data_df, output_df)
		elif spss_test == "mr":
			spss_mr_apa_table(mod_raw_data_df, output_df)
		elif spss_test == "indttest":
			spss_indttest_apa_table(mod_raw_data_df, output_df)
		elif spss_test == "pairttest":
			spss_pairttest_apa_table(mod_raw_data_df, output_df)
	elif input_type == "pvalues":
		pvalues_table(mod_raw_data_df, output_df)


