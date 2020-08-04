import decision_funcs
import global_vars
import helper_funcs
import os

#temp for testing
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from docx import Document
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.enum.table import WD_ALIGN_VERTICAL
from docx.shared import Inches

global_vars.input_path_and_filename = "D:/Desktop_current/MAPAS/GitHub/unit_testing/main_funcs/input dataframes/spss_correlations_tests_pearson.xlsx"
global_vars.alpha_threshold = 0.05
global_vars.output_filename = "D:/Desktop_current/MAPAS/GitHub/a"
global_vars.output_filetype = "Excel"
global_vars.input_type = "spss"
global_vars.raw_test = ""
global_vars.raw_corr_type = ""
global_vars.raw_corr_vars = []
global_vars.raw_corr_include_CI = True
global_vars.raw_mr_outcomevar = ""
global_vars.raw_mr_predictors = []
global_vars.raw_indttest_groupvar = ""
global_vars.raw_indttest_grouplevel1 = ""
global_vars.raw_indttest_grouplevel2 = ""
global_vars.raw_indttest_dv = []
global_vars.raw_pairttest_var_pairs = []
global_vars.summ_corr_varOne = ""
global_vars.summ_corr_varTwo = ""
global_vars.summ_corr_coeff = ""
global_vars.summ_corr_pvalues = ""
global_vars.summ_indttest_var = ""
global_vars.summ_indttest_meanOne = ""
global_vars.summ_indttest_sdOne = ""
global_vars.summ_indttest_nOne = ""
global_vars.summ_indttest_meanTwo = ""
global_vars.summ_indttest_sdTwo = ""
global_vars.summ_indttest_nTwo = ""
global_vars.summ_indttest_equal_var = ""
global_vars.spss_test = "corr"
global_vars.spss_indttest_nOne = -1
global_vars.spss_indttest_nTwo = -1
global_vars.spss_indttest_groupOneLabel = "Group1"
global_vars.spss_indttest_groupTwoLabel = "Group2"
global_vars.spss_pairttest_n = -1
global_vars.pvalues_col = ""
global_vars.effect_size_choice = "no selection"
global_vars.correction_type = "bonferroni"
global_vars.non_numeric_input_raise_errors = False
global_vars.corr_table_triangle = "Both"


def spss_corr_apa_table(mod_raw_data_df, output_df):
	#could very easily use the summ_corr_apa_table function here as I pass identical data - seperate is preferred, however, as might be updated/adjusted in the future
	#note that the raw corr apa table has CIs so not usable
	correlation_label = mod_raw_data_df.iloc[0, 1]
	variables_list = list(mod_raw_data_df.columns)[2:]


	if global_vars.corr_table_triangle == "Upper triangle":
		header_cols = variables_list[1:]
		header_rows = variables_list[:-1]
	elif global_vars.corr_table_triangle == "Lower triangle":
		header_cols = variables_list[:-1]
		header_rows = variables_list[1:]
	elif global_vars.corr_table_triangle == "Both":
		header_cols, header_rows = variables_list, variables_list

	wb = Workbook()
	ws = wb.active

	ws.append([""] + header_cols)
	for ind,var in enumerate(header_rows):
		ws.cell(row=ind+2, column=1).value = var

	inside_loop_ind_start = 2
	for outside_loop_ind in range(2, len(header_rows) + 2):

		if global_vars.corr_table_triangle == "Upper triangle" or global_vars.corr_table_triangle == "Both":
			outside_loop_var = ws.cell(row=outside_loop_ind, column=1).value
		elif global_vars.corr_table_triangle == "Lower triangle":
			outside_loop_var = ws.cell(row=1, column=outside_loop_ind).value
		
		for inside_loop_ind in range(inside_loop_ind_start, len(header_cols) + 2):
			
			if global_vars.corr_table_triangle == "Upper triangle" or global_vars.corr_table_triangle == "Both":
				inside_loop_var = ws.cell(row=1, column=inside_loop_ind).value
			elif global_vars.corr_table_triangle == "Lower triangle":
				inside_loop_var = ws.cell(row=inside_loop_ind, column=1).value

			if outside_loop_var == inside_loop_var:
				ws.cell(row=outside_loop_ind, column=inside_loop_ind).value = 1
			else:
				#here query method is not only slower as it is much smaller dataset but also cannot refer to two different vaiables (colname and val)
				df_filtered = output_df[((output_df["var1"]==outside_loop_var) & (output_df["var2"]==inside_loop_var)) | ((output_df["var1"]==inside_loop_var) & (output_df["var2"]==outside_loop_var))].iloc[0]
				r = df_filtered[correlation_label]
				p = df_filtered["adjusted_pvalues"]
				r = helper_funcs.correlations_format_val(r, p)

				if global_vars.corr_table_triangle == "Upper triangle" or global_vars.corr_table_triangle == "Both":
					ws.cell(row=outside_loop_ind, column=inside_loop_ind).value = r
				elif global_vars.corr_table_triangle == "Lower triangle":
					ws.cell(row=inside_loop_ind, column=outside_loop_ind).value = r
		
		if global_vars.corr_table_triangle != "Both":
			inside_loop_ind_start += 1


	for row in range(1, len(header_rows) + 2):
		for cell in ws[row]:
			cell.alignment = global_vars.alignment_center

	for cell in ws[1]:
		cell.font = global_vars.font_header

	for row in range(2, len(header_rows) + 2):
		ws.cell(row=row, column=1).font = global_vars.font_header

	for cell in ws[len(header_rows) + 1]:
		cell.border = Border(bottom=global_vars.border_APA)

	for cell in ws[1]:
		cell.border = Border(top=global_vars.border_APA, bottom=global_vars.border_APA)

	table_notes = ["**p < 0.01", "*p < {}".format(global_vars.alpha_threshold)]
	table_notes.append("Correlation Coefficient used: {}".format(correlation_label))
	helper_funcs.add_table_notes(ws, table_notes)

	helper_funcs.savefile(wb=wb)

	'''
	start_col = 2
	for row in range(2, len(variables_list)+1):
		row_var = ws.cell(row=row, column=1).value
		for col in range(start_col, len(variables_list)+1):
			col_var = ws.cell(row=1, column=col).value
			#here query method is not only slower as it is much smaller dataset but also cannot refer to two different vaiables (colname and val)
			df_filtered = output_df[((output_df["var1"]==row_var) & (output_df["var2"]==col_var)) | ((output_df["var1"]==col_var) & (output_df["var2"]==row_var))].iloc[0]
			r = df_filtered[correlation_label]
			p = df_filtered["adjusted_pvalues"]
			r = helper_funcs.correlations_format_val(r, p)
			ws.cell(row=row, column=col).value = r
		start_col += 1

	for row in range(1, len(variables_list)+1):
		for cell in ws[row]:
			cell.alignment = global_vars.alignment_center

	for cell in ws[1]:
		cell.font = global_vars.font_header

	for row in range(2, len(variables_list)+1):
		ws.cell(row=row, column=1).font = global_vars.font_header

	for cell in ws[len(variables_list)]:
		cell.border = Border(bottom=global_vars.border_APA)

	for cell in ws[1]:
		cell.border = Border(top=global_vars.border_APA, bottom=global_vars.border_APA)

	table_notes = ["**p < 0.01", "*p < {}".format(global_vars.alpha_threshold)]
	helper_funcs.add_table_notes(ws, table_notes)

	helper_funcs.savefile(wb=wb)
	'''

raw_data_df = decision_funcs.get_raw_data_df()
mod_raw_data_df = decision_funcs.modify_raw_data_df(raw_data_df)
output_df = decision_funcs.generate_output_df(mod_raw_data_df)
output_df = decision_funcs.multitest_correction(output_df)
spss_corr_apa_table(mod_raw_data_df, output_df)
# decision_funcs.save_output(mod_raw_data_df, output_df)

